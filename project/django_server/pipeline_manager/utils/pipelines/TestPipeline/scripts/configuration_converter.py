import openpyxl as opx
from openpyxl import Workbook
import csv
import os
import sys
import shutil

'''
python3 configuration_converter.py INPUT_XLS_FILE OUTPUT_XLS_FILE DATASET.XLS [SHEETS_TO_SKIP]
python3 configuration_converter.py /home/flati/Downloads/rinaldi/ConfigurationsR.xlsx \
        /home/flati/Downloads/rinaldi/ConfigurationsR.converted.xlsx /home/flati/Downloads/dataset.xlsx PRJNA392171,PRJNA341670,Codici
'''

if __name__ == '__main__':
    
    filepath = sys.argv[1]
    output = sys.argv[2]
    dataset_file = sys.argv[3]
    skip = []
    if len(sys.argv) > 4:
        skip = sys.argv[4].split(",")
    
    #if os.path.exists(output_dir):
    #    shutil.rmtree(output_dir)
    
    dataset = {}
    wbook = opx.load_workbook(dataset_file)
    for wsheet in wbook.worksheets:
        sheet_title = wsheet.title.strip().replace(" ", "_")
        dataset[sheet_title] = {}

        row_num = 0
        for row in wsheet.iter_rows():
        
            srr = None
            srx = None
            
            col = 0
            for cell in row:
                if cell.value is not None:
                    if col == 0: srr = cell.value
                    elif col == 1: srx = cell.value
                col += 1
                    
            if srr is not None and srx is not None:
                dataset[sheet_title][srr] = srx
                
            row_num += 1
    
    wb = Workbook()
    wb.remove(wb.active)
    
    wbook = opx.load_workbook(filepath)
    for wsheet in wbook.worksheets:
        sheet_title = wsheet.title.strip().replace(" ", "_")
        if sheet_title in skip:
            print("[SKIP] Analyzing sheet="+str(wsheet.title))
            continue
        
        print("Analyzing sheet="+str(wsheet.title))
#         print(dataset[sheet_title])
        
        ws = wb.create_sheet(sheet_title)
        
        met = set()
        row_num = 0
        for row in wsheet.iter_rows():
            
            row_data = []
            for cell in row:
                if cell.value is not None:
                    row_data.append(cell.value)
                    
            if row_num == 0: 
                row_data = [x.strip().replace(" ", ".") for x in row_data]
                
            add_row = True
            if row_num > 0:
                
                row_data = [x.lower() if "control" in str(x).lower() else x for x in row_data]
                
                srr = row_data[0]
                if srr in dataset[sheet_title]:
                    srx = dataset[sheet_title][srr]
                    row_data[0] = srx
                else:
                    srx = srr
            
                if srx in met: add_row = False
                met.add(srx)
            
            row_num += 1
            
            if add_row:
                ws.append(row_data)
            
    wb.save(output) 
            
            
