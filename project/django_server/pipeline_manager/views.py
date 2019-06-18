from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden
import json
import os
import time
import re
import glob
import datetime
import stat
import shutil
import subprocess
import networkx as nx
import openpyxl as opx

from collections import Counter

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings

import zipfile

from Bio import Entrez
from lxml import etree
from _ast import Num

from pipeline_manager.models import User, Project

from passlib.hash import pbkdf2_sha256
import uuid


# Create your views here.
def login(request):
    data = json.loads(request.body.decode("utf-8"))
#     print(data)
    if "email" not in data: return HttpResponseForbidden(json.dumps({"type": "error", "message": "Please, specify your email."}))
    email = data["email"]
#     
    if "password" not in data: return HttpResponseForbidden(json.dumps({"type": "error", "message": "Please, specify a password."}))
    password = data["password"]

#     email = "tiziano.flati@gmail.com"
#     password = "arc0bal3n0"
    
    try:
        user = User.objects.get(email=email)
        
        if pbkdf2_sha256.verify(password, user.hashed_password):
            response = HttpResponse(json.dumps({"type": "message", "message": "Login successful. Go to <a href='#!/main'>the main page</a>."}))
            response.set_cookie(key='logged_in', value=True)
            response.set_cookie(key='username', value=user.username)
            response.set_cookie(key='is_admin', value=user.is_admin)
            response.set_cookie(key='login_token', value=uuid.uuid4())
            return response
        else:
            return HttpResponseForbidden(json.dumps({"type": "error", "message": "There was an error during the login. Please check your credentials."}))
    
    except User.DoesNotExist:
        return HttpResponseNotFound(json.dumps({"type": "error", "message": "This username is not registered. Please <a href='/register'>register here</a>."}))

    
def logout(request):
    response = HttpResponse(json.dumps({"type": "message", "message": "You logged out successfully. Go to <a href='/main'>the main page</a>."}))

    response.delete_cookie(key='logged_in')
    response.delete_cookie(key='username')
    response.delete_cookie(key='is_admin')
    response.delete_cookie(key='login_token')
    return response


def register(request):
    data = json.loads(request.body.decode("utf-8"))
    print(data)
    
    if "username" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify your username."}), status=403)
    username = data["username"]
    
    if "email" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify a valid e-mail address."}), status=403)         
    email = data["email"]
    
    if "password" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify a password."}), status=403)
    password = data["password"]
    
    if "repassword" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify your password twice."}), status=403)
    repassword = data["repassword"]
    if password != repassword:
        return HttpResponse(json.dumps({"type": "error", "message": "The two entered passwords do not match. Please make sure you entered the very same password."}), status=403)
    
    if "first_name" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify your first name."}), status=403)
    first_name = data["first_name"]
    
    if "last_name" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify your last name."}), status=403)
    last_name = data["last_name"]
    
    if "affiliation" not in data: return HttpResponse(json.dumps({"type": "error", "message": "Please, specify your affiliation."}), status=403)
    affiliation = data["affiliation"]
    
#     privacy_policy_agreement = data["privacy_policy_agreement"] if "privacy_policy_agreement" in data else False 
#     if privacy_policy_agreement is False or privacy_policy_agreement == "ALL":
#         return HttpResponse(json.dumps({"type": "message", "message": "Acceptance of privacy policy is mandatory. Go back to <a href='register' target='_self'>the registration page</a>."}))
    
    try:
        user = User.objects.get(email=email)
        return HttpResponse(json.dumps({"type": "error", "message": "There is already a user registered with this email. Please, use another mail address."}), status=403)
    except User.DoesNotExist:
        print("MAIL CHECK SUCCESSFUL. DOES NOT EXIST.")
        
        try:
            user = User.objects.get(username=username)
            return HttpResponse(json.dumps({"type": "error", "message": "There is already a user registered with this username. Please choose another username."}), status=403)
        except User.DoesNotExist:
            print("USERNAME CHECK SUCCESSFUL. DOES NOT EXIST.")
            print("REGISTERING")            
            hashed_password = pbkdf2_sha256.hash(password)
            user = User(username=username, hashed_password=hashed_password, email=email, first_name=first_name, last_name=last_name, affiliation=affiliation).save()
            return HttpResponse(json.dumps({"type": "message", "message": "Registration is complete. User " + username + " correctly added."}))
        

def get_users_of_project(request, project_id):
    print("GET USERS OF PROJECT", project_id)
    p = Project.objects.get(project_id=project_id)
    
    # Includi il creatore (a parte) (ma non si può rimuovere la condivisione da lui)
    # Includi te stesso (a parte, se amministratore) (ma non si può rimuovere la condivisione da lui)
    
    users = p.users.exclude(username=p.creator.username)
    
    return HttpResponse(serializers.serialize("json", users, use_natural_foreign_keys=True))


def share_project(request, project_id, email):
    print("SHARE PROJECT", project_id, email)
    p = Project.objects.get(project_id=project_id)
    u = User.objects.get(email=email)
    
    if u.is_admin:
        return HttpResponse(json.dumps({"type": "message", "message": "The user " + email + " is an administrator, thus this project is automatically shared with him/her."}), status=409)
    
    if u == p.creator:
        return HttpResponse(json.dumps({"type": "message", "message": "The user " + email + " is the creator of this project, thus it is not possible share this project with him/her."}), status=409)
    
    if p.users.filter(email=email).exists():
        return HttpResponse(json.dumps({"type": "error", "message": "This project was already shared with " + email}), status=409)
    
    p.users.add(u)
    p.save()
    
    return HttpResponse(json.dumps({"type": "message", "message": "Project correctly shared with " + email}))


def remove_share_project(request, project_id, email):
    print("REMOVE SHARE PROJECT", project_id, email)
    p = Project.objects.get(project_id=project_id)
    u = User.objects.get(email=email)
    
    if u.is_admin:
        return HttpResponse(json.dumps({"type": "error", "message": "The user " + email + " is an administrator, thus this project cannot be un-shared with him/her"}), status=409)
    
    if u == p.creator:
        return HttpResponse(json.dumps({"type": "error", "message": "The user " + email + " is the creator of this project, thus this project cannot be un-shared with him/her."}), status=409)
    
    if not p.users.filter(email=email).exists():
        return HttpResponse(json.dumps({"type": "error", "message": "This project was not shared with " + email}), status=409)
    
    p.users.remove(u)
    p.save()
    
    return HttpResponse(json.dumps({"type": "message", "message": "Project correctly un-shared with " + email}))


from django.core import serializers


def users(request):
    print("ALL USERS")
    
    users = User.objects.all()
    users = [x["fields"] for x in json.loads(serializers.serialize("json", users, use_natural_foreign_keys=True))]
    
    for user in users:
        del user['hashed_password']
    
    return HttpResponse(json.dumps(users))

    
def shareable_users(request, project_id):
    print("SHAREABLE USERS OF PROJECT", project_id)
    p = Project.objects.get(project_id=project_id)
    
    users = User.objects.filter(is_admin=False)
    users = users.exclude(email=p.creator.email)
    users = users.difference(p.users.all())
    
    return HttpResponse([x["fields"] for x in serializers.serialize("json", users, use_natural_foreign_keys=True)])


def save_users(request):
    
    users = json.loads(request.body.decode("utf-8"))

    total = 0
    for user in users:
        if "changed" in user and user["changed"]:
            print("Changing user", user)
            u = User.objects.get(email=user["email"])
            for x in user:
                if x != "projects":
                    setattr(u, x, user[x])
            u.save()
            total += 1
            
    return HttpResponse(json.dumps({"message": str(total) + " user" + ("s" if total > 1 else "") + " updated.", "type": "info"}))


def get_cluster_users(request, username, hostname):
#     pipeline = json.loads(request.body.decode("utf-8"))
#     (username, remote_path, hostname) = get_credentials(pipeline)
    
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                           "bash -c 'find $HOME/../../ -maxdepth 2 -type d | sort | uniq | xargs -I {} basename {}'"])
    return HttpResponse(json.dumps(result["message"].split("\n")))


def init_ncbi_tools():
    Entrez.email = "tiziano.flati@gmail.com"
    Entrez.api_key = "ae48c58f9a840e56ee71d28cb464cc988408"


def add_papers(request):
    print(request)
    
    project = json.loads(request.POST.get("project", ""))
    bioproject_id = request.POST.get("bioproject_id", "")
    project_id = project["id"]
    
    bioproject = None
    for p in project["projects"]:
        if p["id"] == bioproject_id:
            bioproject = p
    
    response = []
    for file in request.FILES.values():
        paper_dir = get_paper_dir(project_id)
        
        if not os.path.exists(paper_dir):
            os.makedirs(paper_dir)
            
        filepath = paper_dir + file.name
        default_storage.save(filepath, ContentFile(file.read()))
        
        paper = {
            "url": make_relative(filepath),
            "name": file.name,
            "source": "manual"
        }
        
        bioproject["papers"].append(paper)
        
        response.append(paper)
    
    save_project_raw(project)
    
    return HttpResponse(json.dumps(response))


def delete_paper(request, project_id, paper_name):
    print(request)
    
#     paper_name = os.path.basename(paper_url)
    default_storage.delete(get_paper_dir(project_id) + paper_name)
         
    response = {"type": "info", "message": "File " + paper_name + " correctly deleted."}
         
    return HttpResponse(json.dumps(response))


def split_dataset(dataset):
    
    dataset1 = {}
    dataset2 = {}
    
    for id, item in dataset.items():
        if item["source"] == "manual":
            dataset1[id] = item
        else:
            dataset2[id] = item
    
    return dataset1, dataset2


def merge_datasets(dataset1, dataset2):
    dataset = {}
    
    for k in set(dataset1.keys()).difference(dataset2.keys()):
        dataset[k] = dataset1[k]
        
    for k in set(dataset2.keys()).difference(dataset1.keys()):
        dataset[k] = dataset2[k]
    
    for k in set(dataset1.keys()).intersection(dataset2.keys()):
        if type(dataset1[k]) == int or type(dataset1[k]) == list:
            dataset[k] = dataset1[k] + dataset2[k]
        else:
            dataset[k] = {}
            for x in dataset1[k]:
                dataset[k][x] = dataset1[k][x]
            for x in dataset2[k]:
                dataset[k][x] = dataset2[k][x]
                
    return dataset


def upload_dataset(request):
    print(request)
#     print(request.POST)
#     print(request.POST.get("project", ""))
    
    response = json.loads(request.POST.get("project", ""))
    
    dataset = {}
    
    for file in request.FILES.values():
        filepath = get_temp_dir() + file.name
        print("Analyzing filepath", filepath)
        
        default_storage.save(filepath, ContentFile(file.read()))
        wbook = opx.load_workbook(filepath, read_only=True)
        for wsheet in wbook.worksheets:
            
#             mode = "manual" if wsheet.cell(1, 1).value is not None and wsheet.cell(1, 1).value.lower() == "ids" else "SRA"
#             
#             if mode == "SRA":
#                 for row in wsheet.iter_rows():
#                     for cell in row:
#                         if cell.value is not None and isinstance(cell.value, str) and (cell.value.startswith("PRJNA") or cell.value.startswith("SR") or cell.value.startswith("ER") or cell.value.startswith("DR")):
#                             dataset[cell.value] = 1
#             else:
            header = []
            for row in wsheet.iter_rows():
                sample_id = None
                
                for cell in row:
                    if cell.value is None: continue
                    
                    if cell.row == 1:
                        header.append(cell.value)
                    else:
                        if sample_id is None and cell.column > 3: break # No ID has been provided for this row
                        
                        print(cell.row, cell.column, cell.value)
                        
                        if sample_id is None:
                            if cell.column <= 3 and str(cell.value) != "":
                                sample_id = str(cell.value)
                                dataset[sample_id] = {}
                                h = header[cell.column - 1]
                                dataset[sample_id][h] = str(cell.value)
                                
                                for i in range(cell.column-1):
                                    h = header[i]
                                    dataset[sample_id][h] = ""
                        else:
                            h = header[cell.column - 1]
                            dataset[sample_id][h] = str(cell.value)
                            
    for sample, data in dataset.items():
        print(sample, data)
        data["source"] = "manual" if "path" in data and data["path"] != "" else "SRA"
        data["type"] = "run" if "ids" in data and data["ids"] == sample or "run" in data and data["run"] == sample else "experiment" if data["experiment"] == sample else "bioproject" if data["bioproject"] == sample else None

    # Identify the manual part of the dataset
    manual_raw_dataset, sra_raw_dataset = split_dataset(dataset)
    print("{} MANUAL AND {} SRA".format(len(manual_raw_dataset), len(sra_raw_dataset)))
    
    manual_dataset = {}
    sra_dataset = {}
    
    # Identify the SRA part of the dataset
    if sra_raw_dataset:
        samples = list(sra_raw_dataset.keys())
        print("{} unique SRA IDs found".format(len(samples)))
        
        init_ncbi_tools()
        
        # TODO: Implement repeated fetch
        handle = Entrez.esearch(retmax=100000, db="sra", term=' OR '.join(samples))
        try:
            record = Entrez.read(handle)
        except RuntimeError as e:
            return HttpResponse(json.dumps({"message": str(e)}))
            
        handle.close()
        ids = record["IdList"]
        
        print("{} IDS returned from {} SRA IDS".format(len(ids), len(samples)))
        
        handle2 = Entrez.efetch(db="sra", id=','.join(ids))
        record = handle2.read()
        handle2.close()
        
        tree = etree.fromstring(record)
        convert_ncbi_response(sra_dataset, tree)

        # Filter response with regard to the content of dataset
        filter_dataset(sra_dataset, samples)
        
        add_tags(sra_raw_dataset, sra_dataset)
        propagate_tags(sra_dataset)

    if manual_raw_dataset:
        convert_dataset(manual_dataset, manual_raw_dataset)
        add_tags(manual_raw_dataset, manual_dataset)
        propagate_tags(manual_dataset)

    # Merge datasets
    print("Merging MANUAL and SRA datasets")
    mixed_dataset = merge_datasets(manual_dataset, sra_dataset)
    
    print("Merging BASE project with MIXED dataset")
    final_dataset = merge_datasets(response, mixed_dataset)
                
    # Decorate the response with statistics
    decorate_project(final_dataset)
    
    save_project_raw(final_dataset)
    
    return HttpResponse(json.dumps(final_dataset))


def filter_dataset(response, dataset):
    
    print("DATASET", dataset)
    
    for bioproject in response["projects"]:
        print("Filtering bioproject", bioproject["id"], len(bioproject["experiments"]))
        experiments_to_remove = []
        
        if bioproject["id"] not in dataset:
            
            for experiment in bioproject["experiments"]:
                print("\tFiltering experiment", experiment["id"], experiment["dataset"]["sample_ids"])
                
                runs_to_remove = []
                
                if experiment["id"] not in dataset:
                    for run in experiment["dataset"]["sample_ids"]:
                        if run["id"] not in dataset:
                            runs_to_remove.append(run)
                            
                for run in runs_to_remove:
                    experiment["dataset"]["sample_ids"].remove(run)
                
                if len(experiment["dataset"]["sample_ids"]) == 0:
                    experiments_to_remove.append(experiment)
                
        for experiment in experiments_to_remove:
            bioproject["experiments"].remove(experiment)
            
        print("Final bioproject", bioproject["id"], len(bioproject["experiments"]))

def add_tags(tag_dataset, dataset):
    
    for bioproject in dataset["projects"]:
        add_tag(bioproject["tags"], create_tag("type", "bioproject"))
        add_tag(bioproject["tags"], create_tag("bioproject", bioproject["id"]))
        
        if bioproject["id"] in tag_dataset:
            for tag in extract_tags(tag_dataset[bioproject["id"]]):
                add_tag(bioproject["tags"], tag)
        
        for experiment in bioproject["experiments"]:
            add_tag(experiment["tags"], create_tag("type", "experiment"))
            add_tag(experiment["tags"], create_tag("experiment", experiment["id"]))
        
            if experiment["id"] in tag_dataset:
                for tag in extract_tags(tag_dataset[experiment["id"]]):
                    add_tag(experiment["tags"], tag)
            
            for run in experiment["dataset"]["sample_ids"]:
                add_tag(run["tags"], create_tag("type", "run"))
                add_tag(run["tags"], create_tag("run", run["id"]))
                
                if run["id"] in tag_dataset:
                    for tag in extract_tags(tag_dataset[run["id"]]):
                        add_tag(run["tags"], tag)
                        
def propagate_tags(dataset):
    # Top-down
    for bioproject in dataset["projects"]:
        
        # From bioproject to experiments
        for experiment in bioproject["experiments"]:
            for tag in bioproject["tags"]:
                if tag["type"].lower() not in ["type"]:
                        add_tag(experiment["tags"], tag)
            
            # From experiment to runs           
            for run in experiment["dataset"]["sample_ids"]:
                for tag in experiment["tags"]:
                    if tag["type"].lower() not in ["type"]:
                        add_tag(run["tags"], tag)
                        
    # Bottom-up
    for bioproject in dataset["projects"]:
        
        # From run to experiment
        for experiment in bioproject["experiments"]:
            for run in experiment["dataset"]["sample_ids"]:
                for tag in run["tags"]:
                    if tag["type"].lower() not in ["type", "bioproject", "experiment", "run"]:
                        add_tag(experiment["tags"], tag)
            
            # From experiment to bioproject
            for tag in experiment["tags"]:
                if tag["type"].lower() not in ["type", "bioproject", "experiment", "run"]:
                    add_tag(bioproject["tags"], tag)
                    
def extract_tags(data):
    reserved_keywords = ["ids", "bioproject", "type", "experiment", "run", "path", "path2"]
    
    tags = []
    
    for k in data:
        if k.lower() in reserved_keywords: continue
        if data[k] == "": continue
        
        tags.append(create_tag(k, data[k]))
        
    return tags

def create_tag(type, value):
    return {
        "name": value,
        "type": type.capitalize()
    }
def convert_ncbi_response(dataset, tree):
    
    dataset["projects"] = []
    
    # for each bioproject
    for project_elem in tree.iter("EXPERIMENT_PACKAGE"):
        layout_elem = project_elem.find(".//LIBRARY_LAYOUT")
        layout = layout_elem[0].tag
        
#         print("LAYOUT=", layout)
        experiment_id = project_elem.findtext(".//EXPERIMENT//IDENTIFIERS//PRIMARY_ID")
        bioproject_id = project_elem.findtext(".//STUDY//EXTERNAL_ID[@namespace='BioProject']")
        organism = project_elem.find(".//Pool//Member").attrib["organism"]
#         print("EXPERIMENT_ID=", experiment_id, "BIOPROJECT_ID=", bioproject_id, "ORGANISM=", organism)
        
        attributes = {}
        for attribute in project_elem.iter("SAMPLE_ATTRIBUTE"):
            tag = attribute.findtext("TAG").replace(" ", "_")
            value = attribute.findtext("VALUE")
#             print("SAMPLE ATTRIBUTE", tag, "VALUE", value)
            attributes[tag] = value
        
        platform = project_elem.findtext(".//PLATFORM//INSTRUMENT_MODEL")
        study_title = project_elem.findtext(".//STUDY_TITLE")
        study_type = project_elem.find(".//STUDY_TYPE").attrib["existing_study_type"]
        study_abstract = project_elem.findtext(".//STUDY_ABSTRACT")
        
        paper_id = None
        for el in project_elem.iterfind(".//STUDY_LINK//XREF_LINK"):
            if el.findtext(".//DB") == "pubmed":
                paper_id = el.findtext(".//ID")
                
#         print("paper_id", paper_id)
        
        tags = []
        
        tags.append(create_tag("Platform", platform))
        tags.append(create_tag("Layout", "PE" if layout == "PAIRED" else "SE"))
        if organism and organism is not None: tags.append(create_tag("Organism", organism))
            
        # for each run in the bioproject
        selected_run_ids = []
        total_size = 0
        for run_elem in project_elem.iter("RUN"):
            size = int(run_elem.attrib["size"])
            run_id = run_elem.findtext(".//PRIMARY_ID")
#             print("\t", "RUN_ID=", run_id)
#             print("\t", "SIZE=", size)
#             if samples_set and run_id not in samples_set: continue
            selected_run_ids.append({
                        "type": "run",
                        "id": run_id,
                        "size": size,
                        "tags": [x for x in tags]
                    })
            total_size += size
        
        biosample_id = project_elem.findtext(".//SAMPLE//IDENTIFIERS//EXTERNAL_ID[@namespace=\"BioSample\"]")
        biosample_title = project_elem.findtext(".//SAMPLE//TITLE")
        if biosample_title is None:
            biosample_title = ""
        
#                         "pairedend": layout == "PAIRED",
        bioproject_data = {
            "id": experiment_id,
            "type": "experiment",
            "tags": [x for x in tags],
            "dataset": {
                "size": total_size,
                "bioproject_id": bioproject_id,
                "platform": platform,
                "sample_ids": selected_run_ids,
                "attributes": attributes,
                "biosample_id": biosample_id,
                "biosample_title": biosample_title,
                "study": {
                        "title": study_title,
                        "type": study_type,
                        "abstract": study_abstract
                    }
                }
            }
        
        if organism and organism is not None:
            bioproject_data["dataset"]["genome"] = organism
            
        if paper_id and paper_id is not None:  
            bioproject_data["dataset"]["paper_id"] = paper_id
        
        # Find the bioproject data of this experiment
        result = [x for x in dataset["projects"] if x["id"] == bioproject_id]        
        # If there's none, this is the first time we create the project for this experiment
        if not result:
            empy_project = {
                "id": bioproject_id,
                "type": "bioproject",
                "tags": [],
                "experiments": []
            }
            result = [empy_project]
            dataset["projects"].append(empy_project)

        project = result[0]
        
        project["experiments"].append(bioproject_data)

    
def convert_dataset(response, dataset):

    if "projects" not in response: response["projects"] = []
    
    reserved_keywords = ["ids", "bioproject", "experiment", "path", "path2"]
    
    for sample_id in dataset:
        data = dataset[sample_id]
        
        sample = {
            "type": "run",
            "source": "manual",
            "id": sample_id,
            "tags": [],
            "paths": []
        }

        sample["paths"].append(data["path"])
        if "path2" in data and data["path2"] != "":
            sample["paths"].append(data["path2"])
            
        if len(sample["paths"]) == 1:
            data["layout"] = "SE"
        elif len(sample["paths"]) == 2:
            data["layout"] = "PE"
        else:
            data["layout"] = "UNK"
            
#         genome = ""
        
#         tags = []
#         for k in data:
#             if k in reserved_keywords: continue
#             if k.lower() == "organism":
#                 genome = data[k]
#                 
#             tags.append({
#                 "name": data[k],
#                 "type": k
#             })
        
        # Find the bioproject
        bioproject_id = data["bioproject"] if "bioproject" in data else "Bioproject"
        bioproject = None
        for proj in response["projects"]:
            if proj["id"] == bioproject_id:
                bioproject = proj
                break
        if bioproject is None:
            bioproject = {
                "id": bioproject_id,
                "source": "manual",
                "type": "bioproject",
                "experiments": [],
                "tags": []
            }
            # Add the project
            response["projects"].append(bioproject)

        # Find the experiment
        experiment_id = None
        if "experiment" in data:
            experiment_id = data["experiment"]
        else:
            experiment_id = "Experiment"
#             if genome != "":
#                 experiment_id += "_" + genome
            
        experiment = None
        for exp in bioproject["experiments"]:
            if exp["id"] == experiment_id:
                experiment = exp
                break
            
        if experiment is None: 
            experiment = {
                "id": experiment_id,
                "size": 0,
                "type": "experiment",
                "source": "manual",
                "tags": [],
                "dataset": {
                    "bioproject_id": bioproject_id,
                    "sample_ids": []
                }
            }
            
            # Add the experiment
            bioproject["experiments"].append(experiment)
        
        # Add the run
        experiment["dataset"]["sample_ids"].append(sample)
        

def upload_from_ID_list(request):
    
    print(request)
    
    data = json.loads(request.body.decode("utf-8"))
                        
    init_ncbi_tools()
    
    samples = data["list"].split("\n")
    project = data["project"]
    try:
        p = Project.objects.get(project_id=project["id"])
    except Exception as e:
        print(e)
        return HttpResponse(json.dumps({"message": str(e), "type": "error"}))
    
    db = "sra"
#     prefixes = Counter()
#     if samples:
#         for sample in samples:
#             prefix = sample.replace("[0-9]*$")
#             prefixes[prefix] += 1
#     print(prefixes)
    
#     bioproject_search = True
#     samples_set = set()
#     if samples and (samples[0].startswith("SRR") or samples[0].startswith("DRR") or samples[0].startswith("ERR")):
#         samples_set = set(samples)
    
    print("DB=", db)
    
    try:
        handle = Entrez.esearch(retmax=1000, db=db, term=' OR '.join(samples))
        record = Entrez.read(handle)
        handle.close()
    except RuntimeError as e:
        return HttpResponse(json.dumps({"message": str(e)}))
    
    print("RESULT", record)
    ids = record["IdList"]
    
    print("{}/{} IDS returned".format(len(ids), len(samples)))
    
    handle2 = Entrez.efetch(db=db, id=','.join(ids))
    record = handle2.read()
    handle2.close()
    
#     sra_filepath = get_temp_dir() + "sra.xml"
#     default_storage.save(sra_filepath, ContentFile(record))
    
    tree = etree.fromstring(record)
    
    dataset = {}
    convert_ncbi_response(dataset, tree)

#     found = set()
#     for bioproject in response:
#         found.add(bioproject["id"])
#     print("FOUND", found)
    
#     for x in samples:
#         prefix = re.sub('\d', '', x)
#         
#         if prefix == "PRJNA":
# #             print("NOT_FOUND", x)
# 
#             response.append({
#                     "id": x,
#                     "type": "bioproject",
#                     "experiments": []
#                 })

    # Filter response with regard to the content of dataset
    print("Filtering dataset")
    filter_dataset(dataset, samples)
    
    # Merge datasets
    print("Merging datasets")
    final_dataset = merge_datasets(project, dataset)

    # Decorate the response with statistics
    print("Decorating dataset")
    decorate_project(final_dataset)
    
    # Save the project
#     print("Saving project")
#     save_project_raw(final_dataset)
    
    return HttpResponse(json.dumps(final_dataset))


def modules(request, username, hostname, prefix=None):
    modules = {}

    hostname = "login." + hostname + ".cineca.it"
    scripts = get_util_dir() + "get_modules.sh" + " " + get_util_dir() + "extended_modmap"
    execute_remote_command(["scp", "-i", "pipeline_id_rsa", scripts, username + "@" + hostname + ":"])
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                           "bash -c './" + "get_modules.sh" + "'"])

    last_module = ""
    last_category = ""
    
    for line in result["message"].split("\n"):
        line = line.rstrip()
        
        initial_tabs = len(line) - len(line.lstrip('\t'))
        if initial_tabs == 0:  # profile section
            if line.startswith("Profile: "):
                line = line.replace("Profile: ", "")
                if prefix is None or prefix.lower() in line.lower() or prefix.lower() in "profile":
                    modules[line] = {"label": "profile/" + line}
                
        elif initial_tabs == 2:  # category section
            line = line.lstrip()
            last_category = line
            
        elif initial_tabs == 3:  # specific module section
            line = line.strip("\t")
            if not line.startswith(" "): last_module = line
            elif prefix is None or prefix.lower() in last_module.lower():
                modules[last_category + "#" + last_module] = {"label": last_module, "extra": last_category}
                modules[last_category + "#" + last_module + "/" + line.strip()] = {"label": last_module + "/" + line.strip(), "extra": last_category}

    return HttpResponse(json.dumps(sorted(list(modules.values()), key=lambda k: k['label'])))


def accounts(request, username, hostname, prefix=None):
    accounts = []
    ids = set()
    
    hostname = "login." + hostname + ".cineca.it"
    script = get_util_dir() + "get_accounts.sh"
    execute_remote_command(["scp", "-i", "pipeline_id_rsa", script, username + "@" + hostname + ":"])
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                           "bash -c './" + "get_accounts.sh" + "'"])
    
    print("ACCOUNTS", len(result["message"].split("\n")))
    
    for line in result["message"].split("\n"):
        line = line.strip()
        if prefix is None or prefix in line:
            if line not in ids:
                ids.add(line)
                accounts.append({"label": line})
        
    return HttpResponse(json.dumps(accounts))


def qos(request, username, hostname):
    qos = []
    
    map = {
        "galileo": "gll",
        "marconi_broadwell": "bdw",
        "marconi_knl": "knl",
        "marconi_skl": "skl"
    }
    
    full_hostname = "login." + hostname + ".cineca.it"
    script = get_util_dir() + "get_queues.sh"
    execute_remote_command(["scp", "-i", "pipeline_id_rsa", script, username + "@" + full_hostname + ":"])
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + full_hostname,
                           "bash -c './" + "get_queues.sh" + "'"])
    
    for line in result["message"].split("\n"):
        fields = line.rstrip().split("\t")
        name = fields[0]
        limits = fields[1] if len(fields) > 1 else ""
        if name.startswith(map[hostname]):
            qos.append({"id": name, "name": name, "info": limits})
        
    return HttpResponse(json.dumps(qos))


# def genomes(request, cluster_id):
#     
#     print("ASKED GENOMES OF", cluster_id)
#     
#     genomes = []
# 
#     last_module = ""
#     last_category = ""
#     
#     for line in open(get_util_dir() + "all_modules_" + cluster_id + ".txt"):
#         line = line.rstrip()
#         
#         initial_tabs = len(line) - len(line.lstrip('\t'))
#         
#         if initial_tabs == 2:
#             line = line.lstrip()
#             last_category = line
#             
#         if initial_tabs == 3:
#             line = line.strip("\t")
#             if not line.startswith(" "): last_module = line
#             elif last_category == "data":
#                 
#                 genome = {
#                     "id": last_module + "/" + line.strip(),
#                     "img": "imgs/genomes/genome.png"
#                 }
#                 
#                 if genome["id"] == "ig_Mus_musculus/mm10":
#                     genome["organism"] = "Topo"
#                     genome["img"] = "imgs/genomes/mus_musculus.png"
#                     
#                 elif genome["id"] == "ig_Mus_musculus/mm9":
#                     genome["organism"] = "Topo"
#                     genome["img"] = "imgs/genomes/mus_musculus.png"
#                     
#                 elif genome["id"] == "ig_Rattus_norvegicus/rn6":
#                     genome["organism"] = "Ratto"
#                     genome["img"] = "imgs/genomes/rattus.png"
#                     
#                 elif genome["id"].lower().startswith("ig_homo_") or genome["id"].lower().startswith("homo_"):
#                     genome["organism"] = "Uomo"
#                     genome["img"] = "imgs/genomes/homo_sapiens.png"
#                 
#                 name, version = genome["id"].split("/")
#                 name = name.replace("ig_", "").replace("_", " ")
#                 genome["name"] = name.capitalize() + " (" + version + ")"
#                 
#                 genomes.append(genome)
#     
#     return HttpResponse(json.dumps(genomes))


def projects(request):
    projects = []
    for file in glob.glob(get_projects_dir() + "*.json"):
        project = json.load(open(file))
        if "pipelines" in project: del project["pipelines"]
        if "projects" in project: del project["projects"]
        if "monitor_pipelines" in project: del project["monitor_pipelines"] 
        if "logs" in project: del project["logs"]
        if "counters" in project: del project["counters"]
        
        if "last_modified" not in project:
            project["last_modified"] = project["creation_date"]
        
        projects.append(project)
        
    return HttpResponse(json.dumps(projects))


def pipelines(request):
    return HttpResponse(json.dumps(get_repository_pipelines()))


def get_repository_pipelines():
    pipelines = []
    
    for file in glob.glob(get_util_dir() + "pipelines/*.json"):
        pipeline = json.load(open(file, "r"))
        pipelines.append(pipeline)
        
    return pipelines


def steps(request):
    steps = []
    
    for file in glob.glob(get_util_dir() + "steps/*.json"):
        step = json.load(open(file, "r"))
        if "tags" not in step: step["tags"] = []
        steps.append(step)
    
    return HttpResponse(json.dumps(steps))


def save_pipeline_to_repository(request, overwrite=False):
    overwrite = overwrite in [True, "true", "True"]
    return save_pipeline_to_repository_raw(request, overwrite, get_util_dir() + "pipelines/")


def save_monitor_pipeline_to_repository(request, overwrite=False):
    overwrite = overwrite in [True, "true", "True"]
    return save_pipeline_to_repository_raw(request, overwrite, get_util_dir() + "monitor_pipelines/")

# def public_file_to_local(path):
#     return path.replace("data", )


def save_pipeline_to_repository_raw(request, overwrite, basedir):
    pipeline = json.loads(request.body.decode('utf-8'))
    
    filepath = basedir + pipeline["id"] + ".json"
    
    if os.path.exists(filepath) and not overwrite:
        return HttpResponse(json.dumps({"type": "error", "message": "Pipeline " + pipeline["id"] + " already existing."}))
    else:
        # Save the new file and upload the scripts associated
        json.dump(pipeline, open(filepath, "w"), indent=4, sort_keys=True)
        
        if "files" in pipeline and len(pipeline["files"]) > 0:
            pipeline_dir = basedir + pipeline["id"] + "/scripts/"
            if not os.path.exists(pipeline_dir):
                os.makedirs(pipeline_dir)
                
            for script in pipeline["files"]:
                print("Copying ", get_current_dir() + script["url"], "to", pipeline_dir + script["name"])
                shutil.copy(get_current_dir() + script["url"], pipeline_dir + script["name"])
    
    return HttpResponse(json.dumps({"type": "info", "message": "Pipeline " + pipeline["id"] + " correctly added to the repository."}))


def delete_pipeline_from_repository(request):
    pipeline = json.loads(request.body.decode('utf-8'))
    
    basedir = get_util_dir() + "pipelines/"
    
    filepath = basedir + pipeline["id"] + ".json"
    
    if not os.path.exists(filepath):
        return HttpResponse(json.dumps({"type": "warning", "message": "Pipeline " + pipeline["id"] + " does not exist in the repository, thus it was not removed."}))
    else:
        os.remove(filepath)
        
        if "files" in pipeline and len(pipeline["files"]) > 0:
            pipeline_dir = basedir + pipeline["id"] + "/scripts/"
            if os.path.exists(pipeline_dir):
                shutil.rmtree(pipeline_dir)
                
    return HttpResponse(json.dumps({"type": "info", "message": "Pipeline " + pipeline["id"] + " correctly deleted from the repository."}))


def rename_pipeline_in_repository(request, old_name, new_name):
    
    basedir = get_util_dir() + "pipelines/"
    filepath = basedir + old_name + ".json"
    print("RENAME PIPELINE REPOSITORY", filepath)
    
    if not os.path.exists(filepath):
        return HttpResponse(json.dumps({"type": "warning", "message": "Pipeline " + old_name + " does not exist in the repository and has thus not been renamed."}))
    else:
        old_filepath = filepath
        new_filepath = basedir + new_name + ".json"
        os.rename(filepath, new_filepath)
        
        return HttpResponse(json.dumps({"type": "info", "message": "Pipeline " + old_name + " correctly renamed in the repository to " + new_name + "."}))    


def save_step_to_repository(request, overwrite=False):
    overwrite = overwrite in [True, "true", "True"]
    return save_step_to_repository_raw(request, overwrite, get_util_dir() + "steps/")


def save_monitor_step_to_repository(request, overwrite=False):
    overwrite = overwrite in [True, "true", "True"]
    return save_step_to_repository_raw(request, overwrite, get_util_dir() + "monitor_steps/")


def save_step_to_repository_raw(request, overwrite, basedir):
    print("OVERWRITE", overwrite)
    
    step = json.loads(request.body.decode('utf-8'))

    filepath = basedir + step["title"] + ".json"
    print("ADDING STEP TO REPOSITORY", filepath)
    
    if os.path.exists(filepath) and not overwrite:
        return HttpResponse(json.dumps({"type": "error", "message": "Step " + step["title"] + " already existing."}))
    else:
        # Save the new file
        if not os.path.exists(get_current_dir() + basedir):
            os.makedirs(get_current_dir() + basedir)
            
        json.dump(step, open(filepath, "w"), indent=4, sort_keys=True)
    
    return HttpResponse(json.dumps({"type": "info", "message": "Step " + step["title"] + " correctly added to the repository."}))    


def delete_step_from_repository(request):
    step = json.loads(request.body.decode('utf-8'))
    
    basedir = get_util_dir() + "steps/"
    filepath = basedir + step["title"] + ".json"
    print("DELETE STEP REPOSITORY", filepath)
    
    if not os.path.exists(filepath):
        return HttpResponse(json.dumps({"type": "warning", "message": "Step " + step["title"] + " does not exist in the repository and has thus not been removed."}))
    else:
        os.remove(filepath)
    
    return HttpResponse(json.dumps({"type": "info", "message": "Step " + step["title"] + " correctly removed from the repository."}))    


def rename_step_in_repository(request, old_name, new_name):
    
    basedir = get_util_dir() + "steps/"
    filepath = basedir + old_name + ".json"
    print("RENAME STEP REPOSITORY", filepath)
    
    if not os.path.exists(filepath):
        return HttpResponse(json.dumps({"type": "warning", "message": "Step " + old_name + " does not exist in the repository and has thus not been renamed."}))
    else:
        old_filepath = filepath
        new_filepath = basedir + new_name + ".json"
        os.rename(filepath, new_filepath)
        
        return HttpResponse(json.dumps({"type": "info", "message": "Step " + old_name + " correctly renamed in the repository to " + new_name + "."}))    


def import_pipelines(request):
    data = json.loads(request.body.decode('utf-8'))
    pipelines = data["pipelines"]
    project = data["project"]
    
    for pipeline in pipelines:
        repository_pipelines_dir = get_repository_pipeline_script_dir(pipeline["id"])
        pipeline_script_dir = get_pipeline_script_dir(project["id"], pipeline["id"])
        if not os.path.exists(pipeline_script_dir):
            os.makedirs(pipeline_script_dir)
            
        for file in glob.glob(repository_pipelines_dir + "/*"):
            shutil.copy(file, pipeline_script_dir + os.path.basename(file))
        
        project["pipelines"].append(pipeline)
        
    save_project_raw(project)
    
    response = {"type": "info", "message": str(len(pipelines)) + " pipelines imported successfully."}    
    return HttpResponse(json.dumps(response))

    
def templates(request):
    templates = {}
    
    for file in glob.glob(get_util_dir() + "templates/*.json"):
        template = json.load(open(file, "r"))
        template_name, extension = os.path.splitext(os.path.basename(file))
        templates[template_name] = template
        
    return HttpResponse(json.dumps(templates))


def load_project_from_disk(project_id):
    filepath = get_project_file(project_id)
    
    print("LOADING PROJECT", project_id, "FROM FILEPATH", filepath)
    
    if os.path.exists(filepath):
        project = json.load(open(filepath))
        
        decorate_project(project)
        
        return project
    
    else:
        return None

    
def delete_pipeline(request, project_id, pipeline_id):
    
    script_dir = get_pipeline_script_dir(project_id, pipeline_id)
    if os.path.exists(script_dir):
        shutil.rmtree(script_dir)
        
    project = load_project_from_disk(project_id)
    for pipeline in project["pipelines"]:
        if pipeline["id"] == pipeline_id:
            project["pipelines"].remove(pipeline)
            save_project_raw(project)
            break
        
    return HttpResponse(json.dumps({"type": "info", "message": "Pipeline with ID=" + pipeline_id + " correctly removed."}))


def load_project(request, project_id):
    
    project = load_project_from_disk(project_id)
    
    if project is None:
        return HttpResponse(json.dumps({"type": "error", "message": "It was not possible to load project with ID=" + project_id}))
    else:
        return HttpResponse(json.dumps(project))


def add_tag(tags, tag):
    if tag not in tags:
        tags.append(tag)

def decorate_project(project):
    all_size = "N/A"
    
#     complex_project = "projects" in project
#     projects = project["projects"] if complex_project else project
#     projects = project["projects"]
    
    if "projects" not in project: project["projects"] = []
    
    total_bioprojects = len(project["projects"])
    total_experiments = 0
    total_runs = 0
    counters = {}
    
    for bioproject in project["projects"]:
        
        total_experiments += len(bioproject["experiments"])
        
        size = "N/A"
        for experiment in bioproject["experiments"]:
            
            exp_size = "N/A"
            if "size" in experiment["dataset"]:
                exp_size = experiment["dataset"]["size"]
                if size == "N/A": size = 0
                size += exp_size
            experiment["size"] = exp_size
            
            if "BiosampleId" not in counters: counters["BiosampleId"] = Counter()
            if "biosample_id" in experiment["dataset"]:
                counters["BiosampleId"][experiment["dataset"]["biosample_id"]] += 1
            
            if "Paper" not in counters: counters["Paper"] = Counter()
            if "paper_id" in experiment["dataset"]:
                counters["Paper"][experiment["dataset"]["paper_id"]] += 1
            
            total_runs += len(experiment["dataset"]["sample_ids"])
            
            for run in experiment["dataset"]["sample_ids"]:
                if "tags" in run:
                    for tag in run["tags"]:
                        if tag["type"] not in counters: counters[tag["type"]] = Counter()
                        counters[tag["type"]][tag["name"]] += 1
        
        bioproject["size"] = size
        if all_size == "N/A" and size != "N/A": all_size = 0
        if type(size) == int:
            all_size += size
    
    if "creation_date" not in project:
        project["creation_date"] = str(datetime.datetime.now())
    if "creation_date" in project and "last_modified" not in project:
        project["last_modified"] = project["creation_date"]
    
    project["total_bioprojects"] = total_bioprojects
    project["total_experiments"] = total_experiments
    project["total_runs"] = total_runs
    project["counters"] = counters

    project["size"] = all_size

    if "pipelines" in project:
        for pipeline in project["pipelines"]:
            if "modules" not in pipeline:
                pipeline["modules"] = []
            for step in pipeline["steps"]:
                if "tags" not in step:
                    step["tags"] = []

    
def create_project(request):

    project = json.loads(request.body.decode('utf-8'))
    
    print("NEW PROJECT", project)
        
    project_id = project["id"]
    title = project["title"]
    subtitle = project["subtitle"]
    description = project["description"]
    creator = project["creator"]
    project_path = make_relative(get_project_file(project_id))
    
    project["last_modified"] = project["creation_date"] = str(datetime.datetime.now())
    
    if project_id is "":
        result = {"type": "error", "message": "A project ID must be provided."}
        return HttpResponse(json.dumps(result))
    
    u = User.objects.get(username=creator)

    try:
        p = Project.objects.get(project_id=project_id)
        result = {"type": "error", "message": "A project with id \"" + project_id + "\" already exists"}
    except Project.DoesNotExist:
        
        if os.path.exists(project_path):
            result = {"type": "error", "message": "A project with id \"" + project_id + "\" already exists"}
        else:
            print("Request project does not exist. Creating one.")
            p = Project(project_id=project_id, title=title, subtitle=subtitle, description=description, base_path=project_path, creator=u)
            p.save()
            u.projects.add(p)
            u.save()
            result = {"type": "message", "message": "Correctly created new project with title: " + title + " for user " + creator}
    
            if "img" in project and os.path.exists(get_current_dir() + "/" + project["img"]):
                file_dir = get_file_dir(project["id"])
                if not os.path.exists(file_dir):
                    os.makedirs(file_dir)
                    
                shutil.copy(get_current_dir() + "/" + project["img"], file_dir + os.path.basename(project["img"]))
                
                project["img"] = make_relative(file_dir + os.path.basename(project["img"]))
            
            save_project_raw(project)
    
    return HttpResponse(json.dumps(result))


def save_project(request):
    project = json.loads(request.body.decode('utf-8'))
    save_project_raw(project)
    return HttpResponse(json.dumps({"type": "success", "message": "Project: '{}' correctly saved.".format(project["id"])}))


def save_project_raw(project):
    
    project["last_modified"] = str(datetime.datetime.now())
    
    project_id = project["id"]
    filepath = get_project_file(project_id)
    open(filepath, "w").write(json.dumps(project, indent=4))

    
def delete_project(request, project_id):
    
    if os.path.exists(get_project_file(project_id)):
        os.remove(get_project_file(project_id))
    
    if os.path.exists(get_project_dir(project_id)):
        shutil.rmtree(get_project_dir(project_id))
    
    try:
        p = Project.objects.get(project_id=project_id)
        p.delete()
        result = {"type": "info", "message": "Project: '{}' correctly removed.".format(project_id)}
    except e:
        result = {"type": "error", "message": "Error during the removal of project {}: {}".format(project_id, str(e))}
    
    return HttpResponse(json.dumps(result))

# def rename_project(request):
#     data = json.loads(request.body.decode('utf-8'))
#     
#     project_id = data["id"]
#     new_project_id = data["extra"]
#     filepath = get_project_file(project_id)
#     newfilepath = get_project_file(new_project_id)
#     os.rename(filepath, newfilepath)
#     
#     return HttpResponse("Project: '{}' correctly renamed as '{}'.".format(project_id, new_project_id))


def rename_pipeline(request, project_id, pipeline_id, new_pipeline_id):
    
    project = load_project_from_disk(project_id)
    pipeline = None
    for p in project["pipelines"]:
        if p["id"] == pipeline_id or p["id"] == "" and pipeline_id == "None":
            pipeline = p
            break
    
    if pipeline_id != "":
        old_pipeline_script_dir = get_pipeline_script_dir(project_id, pipeline_id)
        if os.path.exists(old_pipeline_script_dir):
            new_pipeline_script_dir = get_pipeline_script_dir(project_id, new_pipeline_id)
            os.rename(old_pipeline_script_dir, new_pipeline_script_dir)

    pipeline["id"] = new_pipeline_id

    response = {"message": "Pipeline '" + pipeline_id + "' correctly renamed as '" + new_pipeline_id + "'", "type": "info"}
    
    return HttpResponse(json.dumps(response))


def tags_compatibles(g1, g2):
    """
        g2 usually are the pipeline's tags (or step's tags) and 
        g1 are the bioentity's tags
        every type of g2 has to be contained in g1 and the intersection must not be empty
        if g1 is contained in g2 then return compatible
        if g1 intersects g2 then return compatible
    """

    typetag1 = {}
    for x in g1:
        if x["type"] not in typetag1:
            typetag1[x["type"]] = set()
        typetag1[x["type"]].add(x["name"])
        
    typetag2 = {}
    for x in g2:
        if "name" not in x: continue
        
        if x["type"] not in typetag2:
            typetag2[x["type"]] = set()
        typetag2[x["type"]].add(x["name"])
    
    compatible = True
    
    for type in typetag2:
        if type not in typetag1 or len(typetag2[type].intersection(typetag1[type])) == 0:
            compatible = False
            break
    
    return compatible


def get_tags(dataset, bio_entity, level):
    
    tags = list()
    tags_key = set()
    
    if level == "top":
        for project in dataset["projects"]:
            for experiment in project["experiments"]:
                for tag in experiment["tags"]:
                    tag_key = tag["name"] + "#" + tag["type"]
                    if tag_key not in tags_key:
                        tags_key.add(tag_key)
                        tags.append(tag)
    
    if level == "project":
        for project in dataset["projects"]:
            if project["id"] != bio_entity["id"]: continue
            
            for experiment in project["experiments"]:
                for tag in experiment["tags"]:
                    tag_key = tag["name"] + "#" + tag["type"]
                    if tag_key not in tags_key:
                        tags_key.add(tag_key)
                        tags.append(tag)
                    
    if level == "experiment":
        for project in dataset["projects"]:
            for experiment in project["experiments"]:
                if experiment["id"] != bio_entity["id"]: continue
                
                for tag in experiment["tags"]:
                    tag_key = tag["name"] + "#" + tag["type"]
                    if tag_key not in tags_key:
                        tags_key.add(tag_key)
                        tags.append(tag)
                    
    if level == "sample":
        for project in dataset["projects"]:
            for experiment in project["experiments"]:
                samples = [x["id"] for x in experiment["dataset"]["sample_ids"]]
                if bio_entity["id"] not in samples: continue
                
                for tag in experiment["tags"]:
                    tag_key = tag["name"] + "#" + tag["type"]
                    if tag_key not in tags_key:
                        tags_key.add(tag_key)
                        tags.append(tag)
                    
    return tags


levels = ["top", "project", "experiment", "sample"]


def get_command_level_indexes(script_level, command_level):
    script_level_index = levels.index(script_level)
    command_level_index = levels.index(command_level)
    return script_level_index, command_level_index


def produce_script(v, vertex2name, launch, pipeline, step, bioentity):
    if step["skip"]: return None
    
    name = vertex2name[v]
    bioentity_id = bioentity["id"]
    
    script_level = step["script_level"]  # where to put the script
    command_level = step["command_level"]  # what the command is defined over
    command_parallelism_level = step["command_parallelism_level"]  # sequential or parallel
    command_group_level = step["command_group_level"] if "command_group_level" in step else "all"  # all - chunks
    # command_chunk_size = step["command_chunk_size"] #
    
    directives = step["hpc_directives"]
    parallel = step["command_parallelism_level"] == "parallel"
    
    script_level_index, command_level_index = get_command_level_indexes(script_level, command_level)
    
#     levels = ["top", "project", "experiment", "sample"]
#     script_level_index = levels.index(script_level)
#     command_level_index = levels.index(command_level)
    
    sh_file = name + "_" + launch["id"] + ".sh"
    sh_dir = None
    
    bioproject = None
    experiment = None
    
    if script_level_index == 0:  # top
        sh_dir = "./"
        
    if script_level_index == 1:  # project
        bioproject = bioentity
        sh_dir = bioentity_id + "/"
        
    if script_level_index == 2:  # experiment
        bioproject_id = None
        for biopr in launch["projects"]:
            if bioproject_id is not None: break
            for exp in biopr["experiments"]:
                if exp["id"] == bioentity_id:
                    bioproject_id = biopr["id"]
                    bioproject = biopr
                    experiment_id = exp["id"]
                    experiment = bioentity
                    break
        sh_dir = bioproject_id + "/" + bioentity_id + "/"
        
    if script_level_index == 3:  # sample
        bioproject_id = None
        experiment_id = None
        for biopr in launch["projects"]:
            if bioproject_id is not None: break
            for exp in biopr["experiments"]:
                if experiment_id is not None: break
                for sample in exp["dataset"]["sample_ids"]:
                    if sample == bioentity:
                        bioproject_id = biopr["id"]
                        bioproject = biopr
                        experiment_id = exp["id"]
                        experiment = exp
                        break
        # print("SCRIPT BIOENTITY", bioproject_id, experiment_id, bioentity)
        sh_dir = str(bioproject_id) + "/" + str(experiment_id) + "/" + str(bioentity_id) + "/"
    
    script_dir = get_launch_dir(launch["project_id"]) + "data/" + sh_dir
    if not os.path.exists(script_dir): os.makedirs(script_dir)
    filepath = script_dir + sh_file
    
#     print("[INFO] PRODUCE_SCRIPT", filepath, pipeline["id"], step["title"], bioentity_id, script_level, script_level_index)
     
    file = open(filepath, "w")

    file.write("#!/bin/bash\n\n")
    file.write("# Description: {}\n".format(step["description"]))
    file.write("# Short description: {}\n\n".format(step["description_short"]))
    file.write("# Creation time: {}\n\n".format(datetime.datetime.now()))
    
    job_name = directives["job_name"]
    job_name = replace_variables(job_name, launch, bioproject, experiment, pipeline, step, bioentity)
    if not job_name: job_name = step["title"]
    job_name += "-" + str(bioentity_id)
#     job_name = job_name.replace("${PROJECT}", subproject_id)
#     job_name = job_name.replace("${STEP_NAME}", step["title"].replace(" ", "_"))
     
    file.write("#SBATCH --job-name={}\n".format(job_name.replace(" ", "-")))
    if "mpi_procs" in directives and directives["mpi_procs"] is not None and directives["mpi_procs"] > 0:
        file.write("#SBATCH --ntasks={}\n".format(int(directives["mpi_procs"]) * int(directives["nodes"])))
        file.write("#SBATCH --ntasks-per-node={}\n".format(directives["mpi_procs"]))
    else:
        file.write("#SBATCH -N {}\n".format(directives["nodes"]))
        file.write("#SBATCH -n {}\n".format(directives["cpu"]))
    # file.write("#SBATCH -n {}\n".format(directives["mpi_procs"]))
    file.write("#SBATCH -p {}\n".format(directives["queue"] or pipeline["queue"]))
    file.write("#SBATCH --mem={}{}\n".format(directives["memory"]["quantity"], directives["memory"]["size"]))
    file.write("#SBATCH --time {}\n".format(directives["walltime"]))
    file.write("#SBATCH --account {}\n".format(directives["account"] or pipeline["account"]))
     
    if "error" not in directives or directives["error"] == "":
        directives["error"] = step["title"] + ".err"
    directives["error"] = replace_variables(directives["error"], launch, bioproject, experiment, pipeline, step, bioentity).replace(" ", "-")
    file.write("#SBATCH --error {}\n".format(directives["error"]))
     
    if "output" not in directives or directives["output"] == "":
        directives["output"] = step["title"] + ".out"
    directives["output"] = replace_variables(directives["output"], launch, bioproject, experiment, pipeline, step, bioentity).replace(" ", "-")
    file.write("#SBATCH --output {}\n".format(directives["output"]))
     
#     file.write("cd $SLURM_SUBMIT_DIR\n\n")
    
    file.write("\n\n{:#^50}\n{:#^50}\n{:#^50}\n".format("", " Module(s) loading ", ""))
    
    if "genome" in pipeline:
        file.write("\n### Pipeline genome loading\n")
        if pipeline["genome"]["type"] == "path":
            file.write("REFERENCE={}\n".format(pipeline["genome"]["path"]))
    
    if "modules" in pipeline:
        file.write("\n### Pipeline module(s) loading\n")
        
        # Special case of bioinf module
        need_to_load_bioinf = False
        bioinf_already_present = False
        for module in pipeline["modules"]:
            module_label = module if type(module) == str else module["label"]
            if "bioinf" in module_label:
                bioinf_already_present = True
            if "ig_" in module_label or "extra" in module and module["extra"] == "data":
                need_to_load_bioinf = True
        if need_to_load_bioinf and bioinf_already_present:
            file.write("module load autoload profile/bioinf\n")
            
        for module in pipeline["modules"]:
            file.write("module load autoload {}\n".format(module if type(module) == str else module["label"]))
    
    file.write("\n### Step-specific module(s) loading\n")    
    for module in step["modules"]:
        file.write("module load autoload {}\n".format(module if type(module) == str else module["label"]))

    if "begin_block" in step and step["begin_block"]:
        file.write("# Begin block(s)\n")
        begin_block = step["begin_block"]
        begin_block = replace_variables(begin_block, launch, bioproject, experiment, pipeline, step, bioentity)
        file.write("{}\n\n".format(begin_block))
     
    loops = max(0, command_level_index - script_level_index)
     
    if script_level_index == 0:  # top
        
        if command_level_index == 1:
            file.write("""
for PROJECT in `cat projects`; do
    cd $PROJECT
""")

        if command_level_index == 2:
            file.write("""
for PROJECT in `cat projects`; do
    cd $PROJECT
    
    for EXPERIMENT in `cat experiments`; do
        cd $EXPERIMENT
""")

        if command_level_index == 3:
            file.write("""
for PROJECT in `cat projects`; do
    cd $PROJECT
    for EXPERIMENT in `cat experiments`; do
        cd $EXPERIMENT
        for SAMPLE in `cat samples`; do
            cd $SAMPLE
""")
    
    if script_level_index == 1:  # project
        
        if command_level_index == 1:
            file.write("""
PROJECT={}
""".format(bioentity_id))
        
        if command_level_index == 2:
            file.write("""
PROJECT={}
for EXPERIMENT in `cat experiments`; do
    cd $EXPERIMENT
""".format(bioentity_id))
            
        if command_level_index == 3:
            file.write("""
PROJECT={}
for EXPERIMENT in `cat experiments`; do
    cd $EXPERIMENT
    for SAMPLE in `cat samples`; do
        cd $SAMPLE
""".format(bioentity_id))
    
    if script_level_index == 2:  # experiment
        if command_level_index == 2:
            file.write("""
EXPERIMENT={}
""".format(bioentity_id))
        
        if command_level_index == 3:
            file.write("""
    EXPERIMENT={}
    for SAMPLE in `cat samples`; do
        cd $SAMPLE
""".format(bioentity_id))

    if script_level_index == 3:  # run
        file.write("""
PROJECT={}\n
EXPERIMENT={}\n
SAMPLE={}\n
""".format(bioproject["id"], experiment["id"], bioentity_id))
            
    file.write("    " * loops)
    file.write("execute=1 # Overwrite:{}\n".format(step["overwrite"] if "overwrite" in step else ""))

    # Add the run-time conditions to check before launching the command
    if "overwrite" not in step or not step["overwrite"]:
        for condition in step["conditions"]:
            command = condition["command"]
            if command.strip() == "": continue
            
            command = replace_variables(command, launch, bioproject, experiment, pipeline, step, bioentity)
                
            file.write(
                """
                # Checking skip conditions
                execute=0
                {}
                step_condition=$?
                echo "{}"
                echo "{} $step_condition"
                if [ "$step_condition" -eq 1 ]
                then
                    execute=1
                fi
                \n""".format(command, command, name))

    file.write("    " * loops)
    file.write("# Command line(s)\n")
    if "tags" in bioentity:
        file.write("# tags: {}\n".format([t["type"] + (":" + str(t["name"]) if "name" in t else "") for t in bioentity["tags"]]))
    command_line = step["commandline"]
    command_line = replace_variables(command_line, launch, bioproject, experiment, pipeline, step, bioentity)
    file.write("    " * loops)
    file.write("set +o xtrace;\n")
    file.write("    " * loops + "if [ $execute -eq 1 ]; then\n")
    file.write("    " * (loops + 1))
    file.write("{}".format(command_line))
#     if "write_stdout_log" in step and step["write_stdout_log"]:
#         file.write(" >\"${}/{}.out\"".format(sample_variable, step["title"]))
#     if "write_stderr_log" in step and step["write_stderr_log"]:
#         file.write(" 2>\"${}/{}.err\"".format(sample_variable, step["title"]))
    if parallel: file.write(" &\n")
    else: file.write("\n")
    file.write("    " * loops + "fi\n")
     
    while(loops > 0):
        loops -= 1
        file.write("    " * (loops + 1) + "cd ..\n")
        file.write("    " * loops + "done\n")
        
    if parallel:
        file.write("""
wait
        """)
    
    if "end_block" in step and step["end_block"]:
        file.write("# End block(s)\n")
        end_block = step["end_block"]
        end_block = replace_variables(end_block, launch, bioproject, experiment, pipeline, step, bioentity)
        file.write("{}".format(end_block))
    
    file.close()
     
    st = os.stat(filepath)
    os.chmod(filepath, st.st_mode | stat.S_IEXEC)
    
    if "executables" in step:
        for executable in step["executables"]:
            p = script_dir + "/" + executable["filename"]
            f = open(p, "w")
            f.write(executable["command"])
            f.close()
            st = os.stat(p)
            os.chmod(p, st.st_mode | stat.S_IEXEC)
    
    return {"path": filepath, "output": directives["output"], "error": directives["error"]}


def contains(u_bioentity, v_bioentity):
    if u_bioentity["id"] == v_bioentity["id"]: return True
    
    if u_bioentity["type"] == "top":  # top level
        for bioproject in u_bioentity["projects"]:
            if contains(bioproject, v_bioentity):
                return True
        return False
    
    if u_bioentity["type"] == "bioproject":  # project level
        for experiment in u_bioentity["experiments"]:
            if contains(experiment, v_bioentity):
                return True
        return False
    
    if u_bioentity["type"] == "experiment":  # experiment level
        for sample in u_bioentity["dataset"]["sample_ids"]:
            if contains(sample, v_bioentity):
                return True
        return False
    
    else:  # sample level
        return u_bioentity == v_bioentity


def produce_directory_structure(project, script_dir):
    
    # Create project base dir
    if not os.path.exists(script_dir):
        os.makedirs(script_dir)
    
    data_dir = script_dir + "data/"
    
    if os.path.exists(data_dir):
        shutil.rmtree(data_dir)
    
    if not os.path.exists(data_dir):
        os.mkdir(data_dir)
    
    ids = set()
    f = open(data_dir + "projects", "w")
    f.write('\n'.join(["\t".join([x["id"]] + [t["type"] + ":" + t["name"] for t in x["tags"]]) for x in project["projects"] if "disabled" not in x or x["disabled"] is not True]) + "\n")
    f.close()
    for bioproject in project["projects"]:
        if "disabled" not in bioproject or bioproject["disabled"] is not True:
            ids.add(bioproject["id"])
            
        bioproject_dir = data_dir + bioproject["id"] + "/"
        if not os.path.exists(bioproject_dir): os.mkdir(bioproject_dir)
        f = open(bioproject_dir + "experiments", "w")
        f.write('\n'.join(["\t".join([x["id"]] + [t["type"] + ":" + t["name"] for t in x["tags"]]) for x in bioproject["experiments"] if "disabled" not in x or x["disabled"] is not True]) + "\n")
        f.close()
            
        for experiment in bioproject["experiments"]:
            if "disabled" not in experiment or experiment["disabled"] is not True:
                ids.add(experiment["id"])
            experiment_dir = bioproject_dir + experiment["id"] + "/"
            if not os.path.exists(experiment_dir): os.mkdir(experiment_dir)

            if experiment["dataset"]["sample_ids"]:
                f = open(experiment_dir + "samples", "w")
                f.write('\n'.join(["\t".join([str(x["id"])] + [t["type"] + ":" + str(t["name"]) for t in x["tags"]]) for x in experiment["dataset"]["sample_ids"] if "disabled" not in x or x["disabled"] is not True]) + "\n")
                f.close()
            
            for sample in experiment["dataset"]["sample_ids"]:
                if "disabled" not in sample or sample["disabled"] is not True:
                    ids.add(sample["id"])
                sample_dir = experiment_dir + str(sample["id"]) + "/"
                if not os.path.exists(sample_dir): os.mkdir(sample_dir)
    
    # Create file 'dataset' with all the IDs (useful for subsequent filsystem module, for
    # cleaning or moving files around)
    f = open(script_dir + "dataset", "w")
    for id in ids:
        f.write(str(id) + "\n")
    f.close()


def produce_info_file(project, filepath):
    total_experiments = 0
    total_runs = 0
    for bioproject in project["projects"]:
        total_experiments += len(bioproject["experiments"])
        for experiment in bioproject["experiments"]:
            total_runs += len(experiment["dataset"]["sample_ids"])
    
    file = open(filepath, "w")
    file.write("PROJECT_NAME={}\n".format(project["id"]))
    file.write("PROJECT_DESCRIPTION={}\n".format(project["description"]))
    file.write("PROJECT_CREATION_DATE={}\n".format(project["creation_date"]))
    file.write("NUM_BIOPROJECT={}\n".format(len(project["projects"])))
    file.write("NUM_EXPERIMENTS={}\n".format(total_experiments))
    file.write("NUM_RUNS={}\n".format(total_runs))
    file.write("NUM_PIPELINES={}\n".format(len(project["pipelines"])))
    file.close()


import copy


def get_run_id(run):
    return run["id"].split("_")[-1]

    
def create_run(project, launch):
    
    num_runs = max([0] + [int(get_run_id(x)) for x in project["runs"]]) if "runs" in project else 0
    new_id = project["id"] + "_" + launch["id"]
    new_id += "_{:0>3}".format(num_runs + 1)
#     new_id += "_" + str(time.time())
    new_id = new_id.replace(" ", "_").replace(".", "_") 
    
    filtered_project = copy.deepcopy(project)
    filtered_project["id"] = new_id
    filtered_project["type"] = "top"
    filtered_project["project_id"] = project["id"]
    filtered_project["pipelines"] = launch["pipelines"]
    filtered_project["partition"] = launch["partition"]
    filtered_project["parallelism"] = launch["parallelism"]
    filtered_project["tags"] = launch["tags"]
    filtered_project["creation_date"] = str(datetime.datetime.now())
    
    del filtered_project["launches"]
    if "runs" in filtered_project: del filtered_project["runs"]
    if "logs" in filtered_project: del filtered_project["logs"]
    
    chosen = set()
    if "results" not in launch:
        raise Exception("You must select a dataset to create a run")
    
    for x in launch["results"]["matches"]:
        if x["selected"] == True:
            chosen.add(str(x["label"]))
            
    bioprojects_to_remove = list()
    for bioproject in filtered_project["projects"]:
        experiments_to_remove = []
        for experiment in bioproject["experiments"]:
            
            runs_to_remove = []
            for run in experiment["dataset"]["sample_ids"]:
                if str(run["id"]) not in chosen:
                    runs_to_remove.append(run)
                    continue
            for x in runs_to_remove: experiment["dataset"]["sample_ids"].remove(x)
            
            if str(experiment["id"]) not in chosen and len(experiment["dataset"]["sample_ids"]) == 0:
                experiments_to_remove.append(experiment)
        for x in experiments_to_remove: bioproject["experiments"].remove(x)
        
        if str(bioproject["id"]) not in chosen and len(bioproject["experiments"]) == 0:
            bioprojects_to_remove.append(bioproject)
        
    for x in bioprojects_to_remove: filtered_project["projects"].remove(x) 
    
    return filtered_project


def print_dataset(project):
    for bioproject in project["projects"]:
        print(bioproject["id"])
        for experiment in bioproject["experiments"]:
            print("\t" + experiment["id"])
            for sample in experiment["dataset"]["sample_ids"]:
                print("\t\t" + str(sample["id"]))


def get_current_dir():
    return os.path.dirname(__file__)


def make_relative(path):
    return path.replace(get_current_dir() + "/", "")


def get_projects_dir():
    return get_current_dir() + "/data/"


def get_temp_dir():
    return get_current_dir() + "/temp/"


def get_util_dir():
    return get_current_dir() + "/utils/"


def get_project_dir(project_id):
    return get_projects_dir() + project_id + "/"


def get_project_file(project_id):
    return get_projects_dir() + project_id + ".json"


def get_paper_dir(project_id):
    return get_project_dir(project_id) + "papers/"


def get_script_dir(project_id):
    return get_project_dir(project_id) + "scripts/"


def get_pipeline_script_dir(project_id, pipeline_id):
    return get_project_dir(project_id) + "scripts/" + pipeline_id + "/"


def get_repository_pipeline_script_dir(pipeline_id):
    return get_util_dir() + "/pipelines/" + pipeline_id + "/scripts/"


def get_file_dir(project_id):
    return get_project_dir(project_id) + "files/"


def get_launches_dir():
    return get_current_dir() + "/launches/"


def get_launch_dir(project_id):
    return get_launches_dir() + project_id + "/"


def get_base_dir(project_id):
    return get_launch_dir(project_id) + "data/"

    
def get_run_dir(run):
    return get_launches_dir() + run["project_id"] + "/runs/" + run["id"] + "/"


def get_real_run_dir(pipeline, run):
    return os.path.join(pipeline["remote_path"], '') + "runs/" + run["id"] + "/"


def get_real_base_dir(pipeline):
    return os.path.join(pipeline["remote_path"], '') + "data/"


def get_real_file_dir(pipeline):
    return os.path.join(pipeline["remote_path"], '') + "files/"


def get_real_script_dir(pipeline):
    return os.path.join(pipeline["remote_path"], '') + "scripts/" + pipeline["id"] + "/" 


def download_project(request, project_id):
    print("[INFO] Creating archive...")
    archive_path = get_launches_dir() + project_id + '_project.zip'
    if os.path.exists(archive_path): os.remove(archive_path)
    
    print("[INFO] Creating archive ({})...".format(archive_path))
    files = retrieve_file_paths(get_projects_dir() + project_id, project_id)
    with zipfile.ZipFile(archive_path, 'w') as zip_file:
        # write each file seperately
        for file in files:
#             print("Archiving file {} as {} (script_dir={})".format(file, file.replace(script_dir, ""), script_dir))
            zip_file.write(file, arcname=file.replace(get_projects_dir(), ""))
        
        file = get_project_file(project_id)
        zip_file.write(file, arcname=file.replace(get_projects_dir(), ""))

        return HttpResponse(json.dumps(
        {
            "url": "download/" + project_id + "_project.zip",
            "filename": project_id + "_project.zip"
         }))


def produce_scripts(data):
    project = data["project"]
    launch = data["launch"]

    # Produce directory structure
    launch_dir = get_launch_dir(project["id"])
    if not os.path.exists(launch_dir): os.makedirs(launch_dir)
    
    # Produce general directory structure
    produce_directory_structure(project, launch_dir)

    # Produce the run
    run = create_run(project, launch)
    
    print("{:=^50s}".format(" DATASET "))
    print_dataset(run)
    
    # Create run directory
    run_dir = get_run_dir(run)
    if not os.path.exists(run_dir): os.makedirs(run_dir)

    # Produce run-specific directory structure
    produce_directory_structure(run, run_dir)

    # Produce scripts
    produce_pipeline_scripts(run_dir, run, run["pipelines"])
    
    # Produce info file    
    produce_info_file(run, run_dir + "/info")
    
    # Produce job monitor scripts
#     produce_job_monitor_scripts(run_dir + "/job_monitor.sh")
    shutil.copy(get_util_dir() + "job_monitor.sh", run_dir + "job_monitor.sh")
    shutil.copy(get_util_dir() + "monitor.py", launch_dir + "monitor.py")
    shutil.copy(get_util_dir() + "filesystem.py", launch_dir + "filesystem.py")
    
    # File dir
    file_dir = get_file_dir(project["id"])
    if not os.path.exists(file_dir): os.makedirs(file_dir)

    # Create phenodata file
    wb = create_phenodata(project)
    wb.save(file_dir + "phenodata.xlsx")
    # Create dataset file
    wb = create_dataset(project)
    wb.save(file_dir + "dataset.xlsx")
    target_file_dir = launch_dir + "files/"
    if os.path.exists(target_file_dir): shutil.rmtree(target_file_dir)
    shutil.copytree(file_dir, target_file_dir)    
    
    # Script dir
    script_dir = get_script_dir(project["id"])
    if not os.path.exists(script_dir): os.makedirs(script_dir)
    target_script_dir = launch_dir + "scripts/"
    if os.path.exists(target_script_dir): shutil.rmtree(target_script_dir)
    shutil.copytree(script_dir, target_script_dir)
    
    # Create archive
    print("[INFO] Creating archive...")
    archive_basename = get_launches_dir() + run["id"]
    archive_path = archive_basename + '.zip'
    if os.path.exists(archive_path): os.remove(archive_path)
    
    print("[INFO] Creating archive ({})...".format(archive_path))
    files = retrieve_file_paths(launch_dir, run["id"])
    with zipfile.ZipFile(archive_path, 'w') as zip_file:
        # write each file seperately
        for file in files:
#             print("Archiving file {} as {} (script_dir={})".format(file, file.replace(script_dir, ""), script_dir))
            zip_file.write(file, arcname=file.replace(launch_dir, ""))
#     result = shutil.make_archive(archive_basename, 'zip', script_dir)
    print("[INFO] Created archive ", archive_path)
    
    return run


def retrieve_file_paths(dirName, id):
 
  # setup file paths variable
  filePaths = []
   
  # Read all directory, subdirectories and file lists
  for root, directories, files in os.walk(dirName):
      for filename in files:
          # Create the full filepath by using os module.
          filePath = os.path.join(root, filename)
          if filePath.startswith(dirName + "runs/") and not filePath.startswith(dirName + "runs/" + id):
              continue
          
          filePaths.append(filePath)
       
  # return all paths
  return filePaths


def get_compatible_bioentities(project, pipeline, step):
    
    script_level = step["script_level"]  # where to put the script
    # print("COMPATIBLE BIOENTITIES", pipeline["id"], step["title"], script_level)
            
    bio_entities = []
    if script_level == "top":
        bio_entities = [project]
        
    elif script_level == "project":
        for bioproject in project["projects"]:
            if "disabled" in bioproject and bioproject["disabled"] == True: continue
            bio_entities.append(bioproject)
            
    elif script_level == "experiment":
        for bioproject in project["projects"]:
            if "disabled" in bioproject and bioproject["disabled"] == True: continue
            for experiment in bioproject["experiments"]:
                if "disabled" in experiment and experiment["disabled"] == True: continue
                bio_entities.append(experiment)
                
    elif script_level == "sample":
        for bioproject in project["projects"]:
            if "disabled" in bioproject and bioproject["disabled"] == True: continue
            for experiment in bioproject["experiments"]:
                if "disabled" in experiment and experiment["disabled"] == True: continue
                for sample in experiment["dataset"]["sample_ids"]:
                    if "disabled" in sample and sample["disabled"] == True: continue
                    bio_entities.append(sample)
                    
    
    # Take only the compatible bioentities
    compatible_bio_entities = []
    for bio_entity in bio_entities:
        bioentity_tags = get_tags(project, bio_entity, script_level)
        step_tags = (step["tags"] if "tags" in step else []) + (pipeline["tags"] if "tags" in pipeline else [])
        is_compatible = tags_compatibles(bioentity_tags, step_tags)
        
#         print("COMPATIBLE", pipeline["id"], step["title"], script_level, step_tags, bioentity_tags, is_compatible)
        
        if is_compatible:
            compatible_bio_entities.append(bio_entity)

#     print("COMPATIBLE BIOENTITIES AFTER", pipeline["id"], step["title"], script_level, [x["id"] for x in bio_entities])
            
    return compatible_bio_entities


import itertools


def pipelines_to_graph(name2vertex, vertex2name, vertexname2step, vertexname2bioentity, vertex2pipeline, launch, pipelines):
    
    g = nx.DiGraph()
    
    for pipeline in pipelines:
        
        pipeline_id = pipeline["id"]
        if "disabled" in pipeline and pipeline["disabled"]:
            continue
        
        print("=============== {} ==============".format(pipeline_id))
        
        if "remote_path" in pipeline:
            found_basedir = False
            for variable in pipeline["variables"]:
                if variable["key"] == "BASEDIR":
                    found_basedir = True
                    break
            if not found_basedir:
                pipeline["variables"].append({"key": "BASEDIR", "value": pipeline["remote_path"]})
        
        steps = pipeline["steps"]
        
        steps_to_remove = []
        steps_to_add = []
        exploded_steps = {}
        for step in steps:
            if step["skip"]: continue
            
            # In case the step has been tagged with a special tag
            tags_to_split = {}
            if "tags" in step:
                for tag in step["tags"]:
                    if "name" not in tag or tag["name"] == "":
                        tags_to_split[tag["type"]] = set()

            # Take the bioentities associated to this step
            compatible_bio_entities = get_compatible_bioentities(launch, pipeline, step)
#             print("COMPATIBLE", str([x["id"] for x in compatible_bio_entities]))
            for be in compatible_bio_entities:
                if "tags" in be:
                    print("TAGS", be["id"], be["tags"])
                    for tag in be["tags"]:
                        if tag["type"] in tags_to_split:
                            tags_to_split[tag["type"]].add(tag["name"])
#             print("TAGS_TO_SPLIT", tags_to_split)
            
            types = [type for type in tags_to_split]
            types_values = []
            for t in types:
                types_values.append(list(tags_to_split[t]))
#             print("VALUES", types_values)
            
            combinations = []
            if types_values:
                for combination in list(itertools.product(*types_values)):
                    combinations.append([(types[i], c) for i, c in enumerate(combination)])
            
            for combination in combinations:
                suffix = "-".join([k + "=" + v for k, v in combination])
                tag_dict = {}
                for k, v in combination: tag_dict[k] = v
                duplicated_step = copy.deepcopy(step)
                for tag in duplicated_step["tags"]:
                    if tag["type"] in types:
                        tag["name"] = tag_dict[tag["type"]]
                
                duplicated_step["title"] += "_" + suffix
                steps_to_add.append(duplicated_step)
                
                if step["title"] not in exploded_steps: exploded_steps[step["title"]] = set()
                exploded_steps[step["title"]].add(duplicated_step["title"])
            
            if combinations:
                steps_to_remove.append(step)

        for s in steps_to_remove:
            steps.remove(s)            
        for s in steps_to_add:
            steps.append(s)
        
        print("NODES", pipeline["id"], len(g.nodes()))
        
        step2vertices = {}
        stepname2step = {}
        for step in steps:
            step2vertices[step["title"]] = []
            stepname2step[step["title"]] = step
        
        for step in steps:
            if step["skip"]: continue
            
            step_id = step["title"]
            
            if "script_level" not in step:
                print("[WARN] {}-{} has not a script_level defined.".format(pipeline_id, step_id))
                continue
            
            # Take the bioentities associated to this step
            compatible_bio_entities = get_compatible_bioentities(launch, pipeline, step)
            
            # The number of scripts will be the number or required compatible bioentities
            for compatible_bio_entity in compatible_bio_entities:
                bioentity_id = str(compatible_bio_entity["id"])

                vertex_name = pipeline_id + "#" + step_id + "#" + bioentity_id 
                v = vertex_name
                step2vertices[step_id].append(v)
                g.add_node(v)
                name2vertex[vertex_name] = v
                vertex2name[v] = vertex_name
                vertexname2step[vertex_name] = step
                vertexname2bioentity[vertex_name] = compatible_bio_entity
                vertex2pipeline[vertex_name] = pipeline
        
        for step in steps:
            if step["skip"]: continue
            
            step_id = step["title"]
            for dep in step["hpc_directives"]["dependencies"]:
                
                if dep in exploded_steps:
                    dep_extended = exploded_steps[dep]
                else:
                    dep_extended = [dep]
                    
                for d in dep_extended:
                    dep_step = stepname2step[d]
                    if dep_step["skip"]: continue
    
                    dep_id = dep_step["title"]
                    
                    # print("STEP", pipeline_id, step_id, dep_id)
                    
                    for u in step2vertices[dep_id]:
                        u_bioentity = vertexname2bioentity[u]
                        for v in step2vertices[step_id]:
                            v_bioentity = vertexname2bioentity[v]
                            if contains(u_bioentity, v_bioentity) or contains(v_bioentity, u_bioentity):
                                g.add_edge(u, v)
                            
    return g


from collections import OrderedDict


def produce_pipeline_scripts(run_dir, launch, pipelines):
    # Initial graph setup
    name2vertex = {}
    vertex2name = {}
    vertexname2step = {}
    vertexname2bioentity = {}
    vertex2pipeline = {}
    
    # Add the pipelines to the graph
    g = pipelines_to_graph(name2vertex, vertex2name, vertexname2step, vertexname2bioentity, vertex2pipeline, launch, pipelines)

        # Simplify by removing skip nodes
#     vertices_to_remove = set()
#     print("=== VERTICES TO REMOVE ===")
#     for v in vertices_to_remove:
#         print(vertex2name[v])
#     for v in vertices_to_remove:
#         g.remove_node(v)
    
    print("LAUNCH NODES=", len(g.nodes()))
    
    # Optional partitioning into clusters
    clusters = {}
    if "partition" in launch and launch["partition"] == "1":
        
        print("Calculating the clusters")
        
        # Get the tags
        tags = launch["tags"]
        tag_types = [x["type"] for x in tags]
        
        # Identify the clusters
        for v in g.nodes():
            vertex_name = vertex2name[v]
            bio_entity = vertexname2bioentity[vertex_name]
            bio_entity_tags = bio_entity["tags"] if "tags" in bio_entity else []
            
            values = []
            for t in sorted(bio_entity_tags, key=lambda x: x["type"] + "#" + x["name"]):
                if t["type"] in tag_types:
                    values.append(t["name"])
            if len(values) > len(tag_types):
                print("[ERROR] Bioentity has more than 1 value for same tag type:", bio_entity["id"], tag_types, bio_entity_tags)
            key = "#".join(values)
            
#             if not key: key = "external"
            if not key: continue
            
            if key not in clusters: clusters[key] = set()
            clusters[key].add(v)
#         print(len(clusters), " CLUSTERS\n", "\n".join([str((k, len(cluster))) for (k, cluster) in clusters.items()]))
#         for (k, cluster) in clusters.items():
#             print("==== " + str(k))
#             for x in cluster:
#                 print("\t", x)
                
#         external_cluster = clusters.pop("external")
        
        # Clean the clusters
#         additional_clusters = {}
        clusters_to_remove = []
        for k, cluster in clusters.items():
            subgraph = g.subgraph(cluster)
            ccs = nx.connected_components(subgraph.to_undirected())
#             print("CCS", k, len(cluster), sum([1 for cc in ccs]), "connected components")
#             for cc in ccs:
#                 print("\t", "CC", cc)
                
#             print("CCS", k, len(cluster))
            for cc in ccs:
                remove_cc = False
                
#                 print("\t", "CC", cc)
                
                for n in cc:
                    for e in g.in_edges(n):
                        if e[0] not in cluster:
#                             print("REMOVING CC BECAUSE:", e)
                            remove_cc = True
                            break
                    if remove_cc:
                        break
                
                if remove_cc:
                    for n in cc:
                        cluster.remove(n)
#                     additional_clusters[str(k) + "extra"+str(len(additional_clusters))] = cc
                    
            if len(cluster) == 0:
                clusters_to_remove.append(k)
#                         external_cluster.add(n)
        for k in clusters_to_remove:
            del clusters[k]
#         for id,x in additional_clusters.items():
#             clusters[id] = x
#         print(len(clusters), " CLEANED CLUSTERS\n", "\n".join([str((k, len(cluster), str(cluster))) for (k, cluster) in clusters.items()]))
#         for (k, cluster) in clusters.items():
#             print("==== " + str(k))
#             for x in cluster:
#                 print("\t", x)
        
        # Merge clusters together if the user specified a factor
        # In teoria dovrei mettere insieme i cluster solo se non creano cicli nel grafo dei cluster
        parallelism = launch["parallelism"] if "parallelism" in launch else 1
        if parallelism > 1:
            print("[INFO] Collapsing clusters in chunk of size", parallelism)
            clusters_to_collapse = set(clusters.keys())
            
            new_clusters = OrderedDict()
            while len(clusters_to_collapse) > 0:
                v = clusters_to_collapse.pop()
                
#                 print("TRYING NEW BIGGER CLUSTER", v, clusters[v])
                
                new_cluster = clusters[v]
                new_id = v
                
                A = set()
                C = set()
                for x in new_cluster:
                    for e in g.in_edges(x):
                        if e[0] not in new_cluster:
                            A.add(e[0])
                    
                    for e in g.out_edges(x):
                        if e[1] not in new_cluster:
                            C.add(e[1])
                            
#                 print("A="+str(A))
#                 print("C="+str(C))
                    
                merged_clusters = set()
                merged_clusters.add(v)
                for n in clusters_to_collapse:
                    # Try to merge the next element
                    if len(merged_clusters) >= parallelism:
                        break
                    
                    B = set()
                    D = set()
                    for x in clusters[n]:
                        for e in g.in_edges(x):
                            if e[0] not in clusters[n] and e[0] not in new_cluster:
                                B.add(e[0])
                        
                        for e in g.out_edges(x):
                            if e[1] not in clusters[n] and e[1] not in new_cluster:
                                D.add(e[1])
                    
                    take_it = A.isdisjoint(D) and B.isdisjoint(C)
                    
                    if take_it:
#                         print("\tMERGING WITH", n, clusters[n], A, D, B, C)
                        merged_clusters.add(n)
                        new_id += "," + n
                        for x in clusters[n]:
                            new_cluster.add(x)
#                         A = A.union(D)
#                         B = B.union(C)
                        A = A.union(B)
                        C = C.union(D)
#                     else:
#                         print("=== ("+str(len(clusters[n]))+") " + str(clusters[n]))
#                         print("B="+str(B))
#                         print("D="+str(D))
#                         print("A disjoint D = {} ({})".format(A.isdisjoint(D), A.intersection(D)))
#                         print("B disjoint C = {} ({})".format(B.isdisjoint(C), B.intersection(C)))
                
                for x in merged_clusters:
                    if x is not v:
                        clusters_to_collapse.remove(x)
                
                new_clusters[new_id] = new_cluster
                
            clusters = new_clusters
#             print(len(clusters), " NEW CLUSTERS", "\n".join([str((k, len(cluster))) for (k, cluster) in clusters.items()]))
#             for (k, cluster) in clusters.items():
#                 print("==== " + str(k))
#                 for x in cluster:
#                     print("\t", x)
                
        # Add dependencies between chunks and intrachunks
        print("Calculating dependencies from leaves to roots")
        cluster_ids = list(clusters.keys())
        for i in range(len(cluster_ids) - 1):
            id = cluster_ids[i]
            next_id = cluster_ids[i + 1]
            
            # Find leaves of the chunk
            leaves = set()
            cluster = clusters[id]
            for n in cluster:
                is_leaf = True
                for e in g.out_edges(n):
                    if e[1] in cluster:
                        is_leaf = False
                        break
                if is_leaf:
                    leaves.add(n)
                     
            # Find roots of the intrachunk
            roots = set()
            next_cluster = clusters[next_id]
            for n in next_cluster:
                is_root = True
                for e in g.in_edges(n):
                    if e[0] in next_cluster:
                        is_root = False
                        break
                if is_root:
                    roots.add(n)
            
            # Add dependencies between leaves and roots
            for leaf in leaves:
                for root in roots:
                    g.add_edge(leaf, root)
        
    roots = []
    for v in g.nodes():
        if g.in_degree(v) == 0:
            roots.append(v)
    
    # Graph plot
    fake_root = "ROOT"
    g.add_node(fake_root)
    vertex2name[fake_root] = fake_root
    vertexname2step[fake_root] = fake_root
    
    for root in roots:
        g.add_edge(fake_root, root)
        
    print("Producing the DOT graph (and converting into SVG)")
    dot_file = open(run_dir + "/graph.dot", "w")
    dot_file.write("digraph G {\n")
#     print("Subproject", dataset["id"])
    vertices = []
    for v in g.nodes():
        s = vertex2name[v]
        vertices.append(s)
#     print("=== DOT SUMMARY ===")
#     print("V=", vertices)
    
    for i, (k, cluster) in enumerate(clusters.items()):
        v_string = "\n".join(['"' + x + '"' + ";" for x in cluster])
        # print(i, (k, cluster), v_string)
        dot_file.write("""
        
        subgraph cluster_{} {{
            style=filled;
            color=lightgrey;
            node [style=filled, color=green];
            {}
            label = "Cluster #{}={}";
        }}
        
        """.format(i, v_string, i, k))
    
#     for vertex in vertices:
#         dot_file.write("\"{}\"\n".format(vertex))
    edges = []
    for e in g.edges():
        u = vertex2name[e[0]]
        v = vertex2name[e[1]]
        edges.append(u + "->" + v)
        dot_file.write("\"{}\" -> \"{}\";\n".format(u, v))
#     print("E=\n", "\n".join(edges))
    dot_file.write("\n}")
    dot_file.close()
    print("DOT file generated: {}".format(run_dir + "/graph.dot"))
    
#     format = "svg"
#     command = "dot -T{} {}/graph.dot -o {}/graph.{}".format(format, run_dir, run_dir, format)
#     
    # ## Remember that there is native code of networkx
    # #nx.draw(G)
    # #plt.savefig("simple_path.png") # plt comes from: import matplotlib.pyplot as plot
    # #plt.show()
#     print("Converting DOT file into SVG. Command: " + command)
#     ret_code = subprocess.run(command, shell=True)
#     #print("DOT CONVERSION:\nCommand: {}\nReturn code: {}".format(command, ret_code))
#     print("Conversion completed.")
    
    if len(g.nodes()) == 1:
        raise Exception("Error during the generation of project: '{}': no scripts were created!".format(launch["id"]))
    
    # Master script
    print("Creating master script")
    filepath = run_dir + "/" + "run.sh"
    file = open(filepath, "w")
    
    file.write("#!/bin/bash\n\n")
    file.write("# Project ID: {}\n".format(launch["id"]))
    file.write("# Title: {}\n".format(launch["title"]))
    file.write("# Subtitle: {}\n".format(launch["subtitle"]))
    file.write("# Description: {}\n".format(launch["description"]))
    file.write("# Creation time: {}\n\n".format(datetime.datetime.now()))
    file.write("declare -A JOB_IDS\n\n")
    file.write("BASEDIR=`pwd`/../../data/\n")
    file.write("RUNDIR=`pwd`\n")
    
    file.write("cd $BASEDIR")
    
    vertex2script = {}
    for v in g.nodes():
        name = vertex2name[v]
        if name == "ROOT": continue
        
        step = vertexname2step[name]
        bioentity = vertexname2bioentity[name]
        pipeline = vertex2pipeline[name]
        
        # Script production
        script_info = produce_script(v, vertex2name, launch, pipeline, step, bioentity)
        if script_info is None:
            print("[WARNING] Produced a script for a step which has skip = True ({})".format(step["title"]))
        vertex2script[v] = script_info
    
    # Graph BFS-search
    queue = set()
    queue.add(fake_root)
    in_degrees = dict(g.in_degree())
#     print("IN DEGREES", in_degrees)
#     bfs_edges = nx.bfs_edges(g, fake_root)
#     for e in bfs_edges:
    order = []
    while queue:
        v = queue.pop()
        
        if v != fake_root:
            order.append(v)
        
        for u in g.neighbors(v):
            in_degrees[u] -= 1
            
            if in_degrees[u] == 0:
                queue.add(u)
#     print("ORDER", order)
    
    for v in order:
        name = vertex2name[v]
        dependencies = [e_in[0] for e_in in g.in_edges(v) if vertex2name[e_in[0]] is not "ROOT"]
        
        script_info = vertex2script[v]
        script_path = script_info["path"]
        
        # Make path relative
        final_script_path = script_path.replace(get_base_dir(launch["project_id"]), "")
#         print(script_path, final_script_path, get_launch_dir(launch["project_id"]) + "data/")
        
        # Get the bash script simple filename
        sh_name = final_script_path.split("/")[-1]
        
        # Extract the basedir
        basename = '/'.join(final_script_path.split("/")[0:-1])
        
        step = vertexname2step[name]
        bioentity = vertexname2bioentity[name]
        pipeline = vertex2pipeline[name]
        
        real_script_path = get_real_base_dir(pipeline) + final_script_path
        out = get_real_base_dir(pipeline) + basename + "/" + script_info["output"]
        err = get_real_base_dir(pipeline) + basename + "/" + script_info["error"]
#         print("VISITING VERTEX={}".format(vertex2name[v]))
        
        file.write("\n#{}\n#{} {} {}\n#{}\n".format("="*75, "="*25, name, "="*25, "="*75))
        
        if dependencies:
            file.write("DEPS=()\n")
            for dep in dependencies:
                dep_name = vertex2name[dep]
                 
                dep_script = """
    DEP_JOB_NAME="{}"
    echo "DEP JOB NAME="$DEP_JOB_NAME
    DEP_JOB_ID=${{JOB_IDS["$DEP_JOB_NAME"]}}
    echo "DEPJOBID="$DEP_JOB_ID
    if [ ! -z $DEP_JOB_ID ]
    then
        DEPS+=($DEP_JOB_ID)
    fi
    """.format(dep_name)
                 
                file.write(dep_script)
             
            script = """
cd {}
sh_file="{}"
echo "DEPENDENCIES($sh_file)=" ${{DEPS[@]}}
set -o xtrace
if [[ ! -z "$DEPS" ]]
then
    job_long_name=$(sbatch --depend=afterany$(printf ":%s" "${{DEPS[@]}}") ./"$sh_file")
else
    job_long_name=$(sbatch ./"$sh_file")
fi

cd $BASEDIR
 
job_id=$(echo $job_long_name | cut -d' ' -f4)
echo "$sh_file => $job_id"
JOB_IDS["{}"]=$job_id

echo -e $job_id'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}" >> $RUNDIR/job_info

set +o xtrace
""".format(basename, sh_name, name, name, pipeline["id"], bioentity["id"], bioentity["type"], step["title"], real_script_path, out, err)
    
        else:
            script = """
cd {}
sh_file="{}"
set -o xtrace
job_long_name=$(sbatch ./"$sh_file")
job_id=$(echo $job_long_name | cut -d' ' -f4)
echo "$sh_file => $job_id"
cd $BASEDIR
JOB_IDS["{}"]=$job_id

echo -e $job_id'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}"'\\t'"{}" >> $RUNDIR/job_info

set +o xtrace
""".format(basename, sh_name, name, name, pipeline["id"], bioentity["id"], bioentity["type"], step["title"], real_script_path, out, err)
        
        file.write(script)
    
    # Keep track of submitted jobs
    file.write('\nrm $RUNDIR/job_ids; for x in "${JOB_IDS[@]}"; do echo $x >> $RUNDIR/job_ids; done\n')
    file.write('date -I > $RUNDIR/run_info\n')
        
    file.close()
    st = os.stat(filepath)
    os.chmod(filepath, st.st_mode | stat.S_IEXEC)


def replace_variables(x, project, bioproject, experiment, pipeline, step, bioentity):
#     print("REPLACE_VARIABLES", x, step["title"], bioentity["id"])
    
    if bioentity["type"] == "run" and "source" in bioentity and bioentity["source"] == "manual":
        for i, path in enumerate(bioentity["paths"]):
            x = x.replace("${PATHS[" + str(i) + "]}", path)
            x = x.replace("${SIMPLE_PATHS[" + str(i) + "]}", path.split(":")[-1])
    
    run_dir = get_real_run_dir(pipeline, project) + "/data/"
    x = x.replace("${RUN_DIR}", run_dir)
    x = x.replace("$RUN_DIR", run_dir)
    
    file_dir = get_real_file_dir(pipeline)
    x = x.replace("${FILE_DIR}", file_dir)
    x = x.replace("$FILE_DIR", file_dir)
    
    script_dir = get_real_script_dir(pipeline)
    x = x.replace("${SCRIPT_DIR}", script_dir)
    x = x.replace("$SCRIPT_DIR", script_dir)
    
    base_dir = get_real_base_dir(pipeline)
    x = x.replace("${BASE_DIR}", base_dir)
    x = x.replace("$BASE_DIR", base_dir)
    
    run_id = get_run_id(project)
    x = x.replace("${RUN_NUMBER}", run_id)
    x = x.replace("$RUN_NUMBER", run_id)
    
    # BUG: questo sostituiva la stringa EXPERIMENT mi sa (nel TOP LEVEL)
    if "tags" in bioentity:
        for tag in bioentity["tags"]:
            type = tag["type"]
            name = str(tag["name"]) if "name" in tag else None
            if name is None or name == "": continue
            
            x = x.replace("${" + type.upper() + "}", name)
            x = x.replace("$" + type.upper(), name)
    
    # Replace $GENOME[$ORGANISM] with the value of GENOME for that $ORGANISM
    if "enhanced_tags" in pipeline:
        for tag in pipeline["enhanced_tags"]:
            if "columns" in tag:
                for column in tag["columns"]:
                    for v in tag["names"]:
                        variable_name = column + "["+v["name"]+"]"
                        variable_value = v["columns"][column]
                        x = x.replace("${" + variable_name + "}", variable_value)
                        x = x.replace("$" + variable_name, variable_value)
                
    # + [{"key": "USER", "value": pipeline["username"]}]
    for variable in pipeline["variables"]:
        if "key_disabled" in variable and variable["key_disabled"] is True:
            k = variable["key"]
            v = None
            
            if k == "sample_variable":
                v = bioentity["id"]
            if k == "experiment_variable":
                if experiment is not None:
                    v = experiment["id"]
            if k == "project_variable":
                if bioproject is not None:
                    v = bioproject["id"]
            if k == "ALL_SAMPLES":
                v = k
            if k == "cpu_variable":
                v = str(step["hpc_directives"]["cpu"])
            if k == "STEP_NAME":
                v = step["title"]
            
            if v is not None:
                if k in x:
                    x = x.replace("${" + str(k) + "}", v)
                    x = x.replace("$" + str(k), v)
            
        else:
            x = x.replace("${" + variable["key"] + "}", variable["value"])
            x = x.replace("$" + variable["key"], variable["value"])
        
    return x


def get_annotated_objects(project, pipeline, step):
    objects = []
    
    bioentities = get_compatible_bioentities(project, pipeline, step)
    
    basedir = pipeline["remote_path"]
    
    script_level_index, command_level_index = get_command_level_indexes(step["script_level"], step["command_level"])
    
    if command_level_index == 0:
        if project in bioentities:
            objects.append({"basedir": "./", "id": "top"})
    
    if command_level_index == 1:
        for bioproject in project["projects"]:
            if bioproject in bioentities:
                objects.append({"basedir": bioproject["id"], "id": bioproject["id"]})
    
    if command_level_index == 2:
        for bioproject in project["projects"]:
            for experiment in bioproject["experiments"]:
                if experiment in bioentities:
                    objects.append({"basedir": bioproject["id"] + "/" + experiment["id"], "id": experiment["id"]})
    
    if command_level_index == 3:
        for bioproject in project["projects"]:
            for experiment in bioproject["experiments"]:
                for run in experiment["dataset"]["sample_ids"]:
                    if run in bioentities:
                        objects.append({"basedir": bioproject["id"] + "/" + experiment["id"] + "/" + run["id"], "id": run["id"]})
    
    return objects

# from tempfile import NamedTemporaryFile
# def get_monitor_step_data(request):
#     data = json.loads(request.body.decode('utf-8'))
# #     print(data)
#     
#     project = data["project"]
#     pipeline = data["monitor_pipeline"]
#     step = data["step"]
#     
#     files = []
#     if "files" in data:
#         files = data["files"]
#     else:
#         for o in get_annotated_objects(project, pipeline, step):
#             files.append("data/" + o["basedir"] + "/" + step["title"] + "-" + o["id"] + ".json")
#     
# #     print(files)
#     
#     remote_path = pipeline["remote_path"]
#     username = pipeline["username"]
#     hostname = "login." + pipeline["cluster"] + ".cineca.it"
#     step_id = step["title"]
#     script_level = step["script_level"]
#     command_level = step["command_level"]
#     
#     script_level_index, command_level_index = get_command_level_indexes(script_level, command_level)
#     
# #     get_current_dir() + "/temp/monitor_query.txt"
#     tempfile = NamedTemporaryFile()
#     for file in files:
#         tempfile.write((file + "\n").encode())
# 
#     response = []
#     message = execute_remote_command(["scp", "-i", "pipeline_id_rsa", tempfile.name, username+"@"+hostname+":"+remote_path], "Data to be monitored has been sent to the server!")
#     response.append(message)
#     if message["exit_code"] == 0:    
#         message = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
#                                            "bash -c '\
#                                            cd "+remote_path + ";\
#                                            python monitor.py -s "+step_id+" -m multi -i "+step_id+" -l " + os.path.basename(tempfile.name) + "'"])
#         
#         if message["type"] == "success":
#             message["data"] = json.loads(message["message"])
#             message["message"] = "Monitoring data retrieved successfully."
#         
#         response.append(message)
#         
#         message = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
#                                            "bash -c '\
#                                            cd "+remote_path + ";\
#                                            rm " + os.path.basename(tempfile.name) + "'"], "Temporary file ("+os.path.basename(tempfile.name)+ ") correctly removed.")
#         response.append(message)
#     
#     print(response)
#     
#     return HttpResponse(json.dumps(response))

    
def invoke_monitor(request):
    
    run = json.loads(request.body.decode('utf-8'))
    
    messages = []
    data = []
    if "runs" in run:
        print("INVOKING GLOBAL MONITOR")
        partial_messages = []
        temp_data = []
        for r in run["runs"]:
            single_messages, single_data = invoke_monitor_single(r)
            partial_messages += single_messages
            temp_data += single_data
            
        id2messages = {}
        for m in partial_messages:
            if m["message"] not in id2messages:
                id2messages[m["message"]] = []
            id2messages[m["message"]].append(m)
        
        for id, message_list in id2messages.items():
            m = message_list[0]
            m["message"] = str(len(message_list)) + "x " + m["message"]
            messages.append(m)
            
        id2jobs = {}
        for d in temp_data:
            if d["NodeName"] not in id2jobs:
                id2jobs[d["NodeName"]] = []
            id2jobs[d["NodeName"]].append(d)
        
        for id, job_history in id2jobs.items():
            final_job = None
            
            for job in job_history:
                if job["State"]["id"] == "COMPLETED":
                    final_job = job
                    break
            
            if final_job is None:
                last = sorted(job_history, key=lambda x: datetime.datetime.strptime(x["End"], '%Y-%m-%dT%H:%M:%S') if "End" in x and x["End"] != "Unknown" and x["End"] != "N/A" else datetime.datetime.now())[-1]
                final_job = last
            
            data.append(final_job)
            
#         BioentityID: "ExpHomoSapiens"
#             BioentityType: "experiment"
#             Elapsed: "00:00:00"
#             ExitCode: "0:0"
#             JobID: "1161569"
#             JobName: "Unnamed-step-1-ExpHomoSapiens"
#             NodeName: "Unnamed pipeline 1#Unnamed step 1#ExpHomoSapiens"
#             ScriptPath: "/gpfs/scratch/userexternal/tflati00/multigenomes/data/MixedProject/ExpHomoSapiens/Unnamed pipeline 1#Unnamed step 1#ExpHomoSapiens_MultiOrganism_Launch_1_002.sh"
#             StandardError: "/gpfs/scratch/userexternal/tflati00/multigenomes/data/MixedProject/ExpHomoSapiens/Unnamed-step-1.err"
#             StandardOutput: "/gpfs/scratch/userexternal/tflati00/multigenomes/data/MixedProject/ExpHomoSapiens/Unnamed-step-1.out"
#             State: Object { description: "Job has terminated all processes on all nodes with an exit code of zero", id: "COMPLETED", color: "#33cc33", … }
#             StepTitle: "Unnamed step 1"
#             Timelimit: "00:01:00"
#             User: "tflati00"
    else:
        print("INVOKING SINGLE MONITOR")
        messages, data = invoke_monitor_single(run)
    
    return HttpResponse(json.dumps({"messages": messages, "data": data}))

def invoke_monitor_single(run):
    
    job_codes = {}
    job_status_file = get_util_dir() + "job_statuses.txt"
    with open(job_status_file) as f:
        for line in f:
            fields = line.strip().split("\t")
            job_codes[fields[1]] = {"id": fields[1], "code": fields[0], "description": fields[2], "color": fields[3]}
    
    messages = []
    data = []
    
    connections = set()
    for pipeline in run["pipelines"]:
        if "disabled" in pipeline and pipeline["disabled"]:
            continue
        
        connections.add(get_credentials(pipeline))
        
    bioentityID2entityTags = {}
    for bioproject in run["projects"]:
        bioentityID2entityTags[bioproject["id"]] = bioproject["tags"]
        for experiment in bioproject["experiments"]:
            bioentityID2entityTags[experiment["id"]] = experiment["tags"]
            for rr in experiment["dataset"]["sample_ids"]:
                bioentityID2entityTags[str(rr["id"])] = rr["tags"]
        
    for connection in connections:
        username, remote_path, cluster = connection
        run_dir = remote_path + "/runs/" + run["id"]
        
        result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + cluster,
                                           "bash -c 'cd " + run_dir + "; ./job_monitor.sh'"])
        if result["exit_code"] != 0:
            messages.append(result)
            continue
        else:
            print("Formatting job statistics")
            
            lines = result["message"].split("\n")
            
            if len(lines) == 1:
                messages.append({"type": "warning", "exit_code": result["exit_code"], "message": "Monitor data are not available yet...Please try again in a few moments"})
            else:
                messages.append({"type": "success", "exit_code": result["exit_code"], "message": "Monitor data successfully retrieved!"})
                
                header = lines[0].strip().split("\t")
                for line in lines[1:]:
                    fields = line.strip().split("\t")
                    
                    info = {}
                    for i in range(len(header)):
                        k = header[i]
                        v = fields[i] if len(fields) > i else ""
                        info[k] = v
                    
                    info["State"] = job_codes[info["State"].split(" ")[0]]
                    if info["BioentityID"] in bioentityID2entityTags:
                        info["tags"] = bioentityID2entityTags[info["BioentityID"]]
                    
                    data.append(info)
                
    return messages, data

def see_top_lines(request):
    data = json.loads(request.body.decode('utf-8'))
    pipeline = data["pipeline"]
    path = data["path"]
    
    credentials = get_credentials(pipeline)
    (username, remote_path, hostname) = credentials
    
    results = []
    
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                   "bash -c 'cd \"" + remote_path + "\"; head -n 100 \"" + path + "\"'"],
                                   )
    
    if result["exit_code"] == 0:
        results.append({"type": "success", "message": "Path correctly read."})
        result["type"] = "data"
        result["message"] = result["message"]
    results.append(result)
    
    return HttpResponse(json.dumps(results))


def download_file(request):
    data = json.loads(request.body.decode('utf-8'))
    
    data = json.loads(request.body.decode('utf-8'))
    pipeline = data["pipeline"]
    path = data["path"]
    
    credentials = get_credentials(pipeline)
    (username, remote_path, hostname) = credentials
    
    results = []
    
    filename = os.path.basename(path)
    
    result = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                   "bash -c 'cd " + remote_path + "; cat " + path + "'"],
                                   )
    
    if result["exit_code"] == 0:
        response = HttpResponse(result["message"], content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=' + filename
    else:
        results.append(result)
        response = HttpResponse(json.dumps(results))
        
    return response


def download_scripts(request):
    
    data = json.loads(request.body.decode('utf-8'))
    
    run = produce_scripts(data)
        
    return HttpResponse(json.dumps(
        {
            "url": "download/" + run["id"] + ".zip",
            "filename": run["id"] + ".zip"
         }))


def execute_remote_command(COMMAND, messageOK="", ignore_output=False):
    print("Executing command:" + " ".join(COMMAND))
    
    p = subprocess.Popen(COMMAND,
                       shell=False,
                       stdout=subprocess.PIPE,
                       stderr=subprocess.PIPE)
    
    try:
        stdout, stderr = p.communicate()
        ret_code = p.returncode
        if ret_code != 0:
            return {"type": "warning", "exit_code": ret_code, "message": str(stderr.decode("utf-8").strip() + "'")}
        else:
            output = messageOK
            if not ignore_output:
                out = str(stdout.decode("utf-8").strip())
                if out != "":
                    output = out
                    
            return {"type": "success", "exit_code": ret_code, "message": output}
    except Exception as e:
        print(e)
        return {"type": "error", "exit_code":-1, "message": str(e)}


import getpass


def get_credentials(pipeline):
    if "username" not in pipeline:
        return {"type": "error", "message": pipeline["id"] + ": please specify a username"}
    
    if "remote_path" not in pipeline:
        return {"type": "error", "message": pipeline["id"] + ": please specify a directory where to transfer the scripts"}
    
    username = pipeline["username"]
    remote_path = pipeline["remote_path"]
    hostname = "login." + pipeline["cluster"] + ".cineca.it"
    
    return (username, remote_path, hostname)


import traceback


def launch_scripts(request, step):
    data = json.loads(request.body.decode('utf-8'))
    
    step = int(step)
    print("PWD", os.getcwd(), "STEP=", step)
    
    results = []
    
    credentials = set()
    
    launch = data["launch"]
    
    for pipeline in launch["pipelines"]:
        if "disabled" in pipeline and pipeline["disabled"]:
            continue
        
        credential = get_credentials(pipeline)
        if "type" in credential:
            results.append(credential)
            continue
        
        credentials.add(credential)
    
    print("CREDENTIALS", credentials)
    
    status_code = 200
    
    for credential in credentials:
        
        username, remote_path, hostname = credential
        
        if step == 1:
            # Create the remote directory on the cluster
            results.append(execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                                   "bash -c 'mkdir " + remote_path + "'"],
                                                   "Directory " + remote_path + " correctly created."))
        elif step == 2:
            
            project = data["project"]
            try:
                run = produce_scripts(data)
                if "runs" not in project:
                    project["runs"] = []
                project["runs"].append(run)
                
                # Save the project (in order not to lose the created run)
                print("SAVING PROJECT (total runs={})".format(len(project["runs"])))
                save_project_raw(project)
                
                print("EXECUTING RUN ID=" + run["id"])
            
                project_archive_name = run["id"] + ".zip"
                project_archive = get_launch_dir(project_archive_name)
    
                results.append({"type": "data", "exit_code": "200", "message": run})
            
                # Transfer the scripts to the cluster 
                results.append(execute_remote_command(["scp", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", project_archive, username + "@" + hostname + ":" + remote_path],
                                                      "Project files correctly transferred to " + remote_path))
            except Exception as e:
                status_code = 500
                print(traceback.format_exc())
                results.append({"type": "error", "message": str(e)})
                
        elif step == 3:
            run = data["launch"]
            project_archive_name = run["id"] + ".zip"
            
            # Unzipping scripts on cluster
            results.append(execute_remote_command(["ssh", "-i" "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                                   "bash -c 'cd \"" + remote_path + "\"; unzip -o \"" + project_archive_name + "\"'"],
                                                   "Project files correctly unzipped", ignore_output=True))    
        elif step == 4:
            run = data["launch"]
            
            # Launching pipeline on cluster
            results.append(execute_remote_command(["ssh", "-i" "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                                   "bash -c 'cd \"" + remote_path + "\"; cd \"runs/" + run["id"] + "\"; ./run.sh &> run.log'"],
                                                   "Project launched!", ignore_output=True))    
        elif step == 5:
            run = data["launch"]
            
            # Launching job monitor on cluster
            results.append(execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                                                   "bash -c 'cd " + remote_path + "; cd \"runs/" + run["id"] + "\"; ./job_monitor.sh'"],
                                                   "Job monitor launched!", ignore_output=True))        
    
    print("STEP", step)
    # status_code = 201 if step < 5 else 200
    
    response = HttpResponse(json.dumps(results))
    response.status_code = status_code
    
    return response


def filesystem_api(request, op):
    
    print("FS API", os.getcwd(), "OP=", op)
    
    results = []
    
    data = json.loads(request.body.decode('utf-8'))
    
    if "results" not in data:
        print(data)
        
        if "ci" not in data:
            results.append({"type": "error", "message": "Please specify connection details"})
        else:
            if "username" not in data["ci"]:
                results.append({"type": "error", "message": "Please specify a username"})
        
            if "remote_path" not in data["ci"]:
                results.append({"type": "error", "message": "Please specify a directory to scan"})
        
            basedir = ""
            if "basedir" in data:
                basedir = data["basedir"]
            
            if not results:
                (username, remote_path, hostname) = get_credentials(data["ci"])
                
        #         command = "bash -c 'find "+remote_path+" -maxdepth 2 -name \'*PRJ*\''"
                command = "cd " + remote_path + "; python filesystem.py " + basedir
                result = execute_remote_command(["ssh",
                                       "-i", "pipeline_id_rsa",
                                       "-oPasswordAuthentication=no",
                                       username + "@" + hostname,
                                       command])
                
                if result["exit_code"] == 0:
                    result["message"] = json.loads(result["message"])
                    result["type"] = "data"
            
                results.append(result)
        
    else:
        print("CHANGED", data["changed"])
        
        changed = data["changed"]
        
        # Treat the special case of indeterminate case
        # From "indeterminate" we can only go to "selected=false"
        if changed["indeterminate"]:
            for group in data["results"]["filters"]:
                
                if group == changed:
                    group["indeterminate"] = changed["indeterminate"] = False
                    group["selected"] = changed["selected"] = False
                
                for filter in group["values"]:
                    if filter == changed:
                        filter["indeterminate"] = changed["indeterminate"] = False
                        filter["selected"] = changed["selected"] = False
        
        for m in data["results"]["matches"]:
            
            remove_it = False
            vote = True
            
            for group in data["results"]["filters"]:
                if not vote:
                    break
                
                for filter in group["values"]:
                    if m[group["label"]] == filter["value"]:
                        
                        if filter == changed or group == changed:
                            remove_it = not changed["selected"]
#                             print("Treating changed filter:", group["label"], filter["value"], m["run"], remove_it)
                            vote = False
                            break
                        
                        else:
                            if not filter["indeterminate"]:
                                remove_it |= not filter["selected"]
                            
            m["selected"] = not remove_it
            
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                filter["partial_count"] = 0
                
        for m in data["results"]["matches"]:
            if m["selected"]:
                for group in data["results"]["filters"]:
                    for filter in group["values"]:
                        if m[group["label"]] == filter["value"]:
                            filter["partial_count"] += 1
        
        # Update filter groups' info
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                if filter["partial_count"] == filter["count"]:
                    filter["indeterminate"] = False
                    filter["selected"] = True
                    
                elif filter["partial_count"] > 0 and filter["partial_count"] < filter["count"]:
                    filter["indeterminate"] = True
                    filter["selected"] = False
                
                elif filter["partial_count"] == 0:
                    filter["indeterminate"] = False
                    filter["selected"] = False
                
                else:
                    filter["indeterminate"] = False
                    filter["selected"] = False
        
            num_selected = sum([1 for x in group["values"] if x["selected"]])
            num_indeterminate = sum([1 for x in group["values"] if x["indeterminate"]])
            group["indeterminate"] = num_indeterminate > 0 or num_selected > 0 and num_selected < len(group["values"])
            
            if num_selected == len(group["values"]): group["selected"] = True
            elif num_selected == 0: group["selected"] = False
        
#         print("{} elements to remove".format(len(to_remove)))
#         for path,m in to_remove.items():
#             data["results"]["matches"].remove(m)
        
        results = [{"message": data["results"], "type": "data"}]
    
#     print(results)
    
    return HttpResponse(json.dumps(results))


def to_human_size(num):
    if type(num) != int: return num
    
    for unit in ['B', 'KB', 'MB', 'GB', 'TB', 'PB']:
        if abs(num) < 1024.0:
            return "%3.1f%s" % (num, unit)
        num /= 1024.0

        
def job_search_api(request):        
    print("JOB SEARCH API")
    
    results = []
    data = json.loads(request.body.decode('utf-8'))
    if "results" not in data:

        matches = []
        filters = []
        
        response = {
            "matches": matches,
            "filters": filters,
            "options": {}
        }
 
        for job in data:
            job["tags"] += [
                    create_tag("ExitCode", job["ExitCode"]),
                    create_tag("StepTitle", job["StepTitle"]),
                    create_tag("State", job["State"]["id"])
                ]
            matches.append({
                "data": job,
                "selected": True,
                "tags": job["tags"]
            })
        
        tags = {}
        for job in matches:
            for tag in job["tags"]:
                tag_name = tag["type"]
                tag_value = tag["name"]
                if tag_name not in tags: tags[tag_name] = Counter()
                tags[tag_name][tag_value] += 1
            
        for tag_name, tag_counter in tags.items():
            filters.append({
                "label": tag_name,
                "selected": True,
                "indeterminate": False,
                "values": [{"value": value, "count": count, "partial_count": count, "selected": True, "indeterminate": False} for value, count in tag_counter.items()]
            })
        
        results.append({"message": response})
        
    else:
        print("CHANGED", data["changed"])
        
        changed = data["changed"]
        
        # Treat the special case of indeterminate case
        # From "indeterminate" we can only go to "selected=false"
        if changed["indeterminate"]:
            for group in data["results"]["filters"]:
                
                if group == changed:
                    group["indeterminate"] = changed["indeterminate"] = False
                    group["selected"] = changed["selected"] = False
                
                for filter in group["values"]:
                    if filter == changed:
                        filter["indeterminate"] = changed["indeterminate"] = False
                        filter["selected"] = changed["selected"] = False
        
        for m in data["results"]["matches"]:
            
            remove_it = False
            vote = True
            
            for group in data["results"]["filters"]:
                if not vote:
                    break
                
                for filter in group["values"]:

                    compatible = False                    
                    for tag in m["tags"]:
                        if tag["type"] == group["label"] and tag["name"] == filter["value"]:
                            compatible = True
                            
                    if compatible:
                        
                        if filter == changed or group == changed:
                            remove_it = not changed["selected"]
#                             print("Treating changed filter:", group["label"], filter["value"], m["run"], remove_it)
                            vote = False
                            break
                        
                        else:
                            if not filter["indeterminate"]:
                                remove_it |= not filter["selected"]
                            
            m["selected"] = not remove_it
        
#         matches_id_to_objects = {}
#         for m in data["results"]["matches"]:
#             matches_id_to_objects[m["label"]] = m
#         
#         # Reconsider containment property
#         for bioproject in data["projects"]:
#             experiments_selected = 0
#             for experiment in bioproject["experiments"]:
#                 runs_selected = 0
#                 for run in experiment["dataset"]["sample_ids"]:
#                     if matches_id_to_objects[run["id"]]["selected"]:
#                         runs_selected += 1
#                 if runs_selected > 0:
#                     matches_id_to_objects[experiment["id"]]["selected"] = True
#                 if matches_id_to_objects[experiment["id"]]["selected"]:
#                     experiments_selected += 1
#             if experiments_selected > 0:
#                 matches_id_to_objects[bioproject["id"]]["selected"] = True
                    
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                filter["partial_count"] = 0
                
        for m in data["results"]["matches"]:
            if m["selected"]:
                for group in data["results"]["filters"]:
                    for filter in group["values"]:
                        compatible = False                    
                        for tag in m["tags"]:
                            if tag["type"] == group["label"] and tag["name"] == filter["value"]:
                                compatible = True
                        if compatible:
                            filter["partial_count"] += 1
        
        # Update filter groups' info
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                if filter["partial_count"] == filter["count"]:
                    filter["indeterminate"] = False
                    filter["selected"] = True
                    
                elif filter["partial_count"] > 0 and filter["partial_count"] < filter["count"]:
                    filter["indeterminate"] = True
                    filter["selected"] = False
                
                elif filter["partial_count"] == 0:
                    filter["indeterminate"] = False
                    filter["selected"] = False
                
                else:
                    filter["indeterminate"] = False
                    filter["selected"] = False
        
            num_selected = sum([1 for x in group["values"] if x["selected"]])
            num_indeterminate = sum([1 for x in group["values"] if x["indeterminate"]])
            group["indeterminate"] = num_indeterminate > 0 or num_selected > 0 and num_selected < len(group["values"])
            
            if num_selected == len(group["values"]): group["selected"] = True
            elif num_selected == 0: group["selected"] = False
        
#         print("{} elements to remove".format(len(to_remove)))
#         for path,m in to_remove.items():
#             data["results"]["matches"].remove(m)
        
        results = [{"message": data["results"]}]
    
#     print(results)
    
    return HttpResponse(json.dumps(results))


def create_new_launch(request, project_id):
    data = json.loads(request.body.decode('utf-8'))
    
    project = load_project_from_disk(project_id)
    
    # Create a new default launch
    launch = json.load(open(get_util_dir() + "templates/launch.json", "r"))
    launch["id"] = "Launch " + str(len(project["launches"]) + 1)
    launch["title"] = project["id"]
    launch["subtitle"] = project["subtitle"]
    launch["description"] = project["description"];
    launch["creation_date"] = str(datetime.datetime.now());
    launch["pipelines"] = copy.deepcopy(project["pipelines"]);
    
    filters = data["filters"]
    matches = data["matches"]
    
    # Identify the steps to enable
    for pipeline in launch["pipelines"]:
        for step in pipeline["steps"]:
            step["skip"] = True
            
    for pipeline in launch["pipelines"]:
        for step in pipeline["steps"]:
            if not step["skip"]: continue
            
            for match in matches:
                if not match["selected"]: continue
                
                match_pipeline_id = match["data"]["Pipeline"] if "pipeline" in match["data"] else None
                if match_pipeline_id is not None and match_pipeline_id != pipeline["id"]: continue
                
                match_step_id = match["data"]["StepTitle"]
                if match_step_id != step["title"]: continue
                
                set_pipeline_skip_step(pipeline, step, False)
                break
            
    # Identify the entities
    chosen = {}
    for match in matches:
        if match["data"]["BioentityID"] not in chosen: chosen[match["data"]["BioentityID"]] = copy.deepcopy(match)
        chosen[match["data"]["BioentityID"]]["selected"] |= match["selected"]

    launch["results"] = {"filters": filters, "matches": [
        {
            "label": x,
            "selected": item["selected"],
            "tags": item["tags"],
            "type": item["data"]["BioentityType"]} for x, item in chosen.items()]}
    
    return HttpResponse(json.dumps(launch))


def set_pipeline_skip_step(pipeline, step, value):
    step["skip"] = value
    
    print("Setting step", step["title"], " to skip=", step["skip"])
    
    for s in pipeline["steps"]:
        if step["title"] in s["hpc_directives"]["dependencies"]:
            set_pipeline_skip_step(pipeline, s, value)


def dataset_api(request):
    
    print("DATASET API")
    
    results = []
    
    data = json.loads(request.body.decode('utf-8'))
    
    if "projects" in data:
        decorate_project(data)
    
    if "results" not in data:

        matches = []
        filters = []
        
        response = {
            "matches": matches,
            "filters": filters,
            "options": {}
        }
                
        for bioproject in data["projects"]:
            matches.append({
                    "selected": True,
                    "type": "bioproject",
                    "label": bioproject["id"],
                    "size": bioproject["size"],
                    "human_size": to_human_size(bioproject["size"]),
                    "tags": bioproject["tags"]
                })
            
            for experiment in bioproject["experiments"]:
                matches.append({
                    "selected": True,
                    "type": "experiment",
                    "label": experiment["id"],
                    "size": experiment["size"],
                    "human_size": to_human_size(experiment["size"]),
                    "tags": experiment["tags"]
                })
                
                for run in experiment["dataset"]["sample_ids"]:
                    matches.append({
                        "selected": True,
                        "type": "run",
                        "label": run["id"],
                        "size": run["size"] if "size" in run else "N/A",
                        "human_size": to_human_size(run["size"]) if "size" in run else "N/A",
                        "tags": run["tags"]
                    })
            
        tags = {}
        raw_tags = []
        for bioproject in data["projects"]:
            raw_tags += bioproject["tags"]
            for experiment in bioproject["experiments"]:
                raw_tags += experiment["tags"]
                for run in experiment["dataset"]["sample_ids"]:
                    raw_tags += run["tags"]
        for tag in raw_tags:
            tag_name = tag["type"]
            tag_value = tag["name"]
            if tag_name not in tags: tags[tag_name] = Counter()
            tags[tag_name][tag_value] += 1
            
        for tag_name, tag_counter in tags.items():
            filters.append({
                "label": tag_name,
                "selected": True,
                "indeterminate": False,
                "values": [{"value": value, "count": count, "partial_count": count, "selected": True, "indeterminate": False} for value, count in tag_counter.items()]
            })
        
        results.append({"message": response})
        
    else:
        print("CHANGED", data["changed"])
        
        changed = data["changed"]
        
        # Treat the special case of indeterminate case
        # From "indeterminate" we can only go to "selected=false"
        if changed["indeterminate"]:
            for group in data["results"]["filters"]:
                
                if group == changed:
                    group["indeterminate"] = changed["indeterminate"] = False
                    group["selected"] = changed["selected"] = False
                
                for filter in group["values"]:
                    if filter == changed:
                        filter["indeterminate"] = changed["indeterminate"] = False
                        filter["selected"] = changed["selected"] = False
        
        for m in data["results"]["matches"]:
            
            remove_it = False
            vote = True
            
            for group in data["results"]["filters"]:
                if not vote:
                    break
                
                for filter in group["values"]:

                    compatible = False                    
                    for tag in m["tags"]:
                        if tag["type"] == group["label"] and tag["name"] == filter["value"]:
                            compatible = True
                            
                    if compatible:
                        
                        if filter == changed or group == changed:
                            remove_it = not changed["selected"]
#                             print("Treating changed filter:", group["label"], filter["value"], m["run"], remove_it)
                            vote = False
                            break
                        
                        else:
                            if not filter["indeterminate"]:
                                remove_it |= not filter["selected"]
                            
            m["selected"] = not remove_it
        
        matches_id_to_objects = {}
        for m in data["results"]["matches"]:
            matches_id_to_objects[m["label"]] = m
        
        # Reconsider containment property
        for bioproject in data["projects"]:
            experiments_selected = 0
            for experiment in bioproject["experiments"]:
                runs_selected = 0
                for run in experiment["dataset"]["sample_ids"]:
                    if matches_id_to_objects[run["id"]]["selected"]:
                        runs_selected += 1
                if runs_selected > 0:
                    matches_id_to_objects[experiment["id"]]["selected"] = True
                if matches_id_to_objects[experiment["id"]]["selected"]:
                    experiments_selected += 1
            if experiments_selected > 0:
                matches_id_to_objects[bioproject["id"]]["selected"] = True
                    
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                filter["partial_count"] = 0
                
        for m in data["results"]["matches"]:
#             if m["type"] is not "experiment": continue
            
            if m["selected"]:
                for group in data["results"]["filters"]:
                    for filter in group["values"]:
                        compatible = False                    
                        for tag in m["tags"]:
                            if tag["type"] == group["label"] and tag["name"] == filter["value"]:
                                compatible = True
                        if compatible:
                            filter["partial_count"] += 1
        
        # Update filter groups' info
        for group in data["results"]["filters"]:
            for filter in group["values"]:
                if filter["partial_count"] == filter["count"]:
                    filter["indeterminate"] = False
                    filter["selected"] = True
                    
                elif filter["partial_count"] > 0 and filter["partial_count"] < filter["count"]:
                    filter["indeterminate"] = True
                    filter["selected"] = False
                
                elif filter["partial_count"] == 0:
                    filter["indeterminate"] = False
                    filter["selected"] = False
                
                else:
                    filter["indeterminate"] = False
                    filter["selected"] = False
        
            num_selected = sum([1 for x in group["values"] if x["selected"]])
            num_indeterminate = sum([1 for x in group["values"] if x["indeterminate"]])
            group["indeterminate"] = num_indeterminate > 0 or num_selected > 0 and num_selected < len(group["values"])
            
            if num_selected == len(group["values"]): group["selected"] = True
            elif num_selected == 0: group["selected"] = False
        
#         print("{} elements to remove".format(len(to_remove)))
#         for path,m in to_remove.items():
#             data["results"]["matches"].remove(m)
        
        results = [{"message": data["results"]}]
    
#     print(results)
    
    return HttpResponse(json.dumps(results))

        
def launch_monitor_scripts(request):
    project = json.loads(request.body.decode('utf-8'))
    
    print("PWD", os.getcwd())
    
    produce_scripts(request)
    
    results = []
    
    print("USER", getpass.getuser())
    
    for pipeline in project["monitor_pipelines"]:
        if "disabled" in pipeline and pipeline["disabled"]:
            continue
        
        if "username" not in pipeline:
            results.append({"type": "error", "message": "Please specify a username"})
            continue
        if "remote_path" not in pipeline:
            results.append({"type": "error", "message": "Please specify a directory where to transfer the scripts"})
            continue
        
        username = pipeline["username"]
        remote_path = pipeline["remote_path"]
        project_archive_name = project["id"] + ".zip"
        project_archive = get_launch_dir(project_archive_name)
        hostname = "login." + pipeline["cluster"] + ".cineca.it"
        
        results.append(execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname, "bash -c 'mkdir " + remote_path + "'"], "Directory " + remote_path + " correctly created."))
        if results[-1]["exit_code"] == 0 or results[-1]["exit_code"] == 1:
            results.append(execute_remote_command(["scp", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", project_archive, username + "@" + hostname + ":" + remote_path], "Project files correctly transferred to " + remote_path))
        if results[-1]["exit_code"] == 0:
            results.append(execute_remote_command(["ssh", "-i" "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname, "bash -c 'cd " + remote_path + "; unzip -o " + project_archive_name + "'"], "Project files correctly unzipped", ignore_output=True))    
        if results[-1]["exit_code"] == 0:
            results.append(execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname, "bash -c 'cd " + remote_path + "; cd data; ./monitor.sh'"], "Monitor launched!", ignore_output=True))        
    
    print(results)
    
    return HttpResponse(json.dumps(results))


import csv


def upload_phenodata(request, project_id):
    
    print(request)
#     project = load_project_from_disk(project_id)
#     script_dir = produce_directory_structure(project)
    
    print(request.FILES)
    for file in request.FILES.values():
        phenodata_dir = get_project_dir(project_id)
        filepath = phenodata_dir + "phenodata.xlsx"
        
        if not os.path.exists(phenodata_dir):
            os.makedirs(phenodata_dir)
        
        # Save the xlsx
        default_storage.save(filepath, ContentFile(file.read()))
        
#         # Generate the phenodata for each bioproject
#         wbook = opx.load_workbook(filepath)
#         for wsheet in wbook.worksheets:
#             sheet_title = wsheet.title.strip().replace(" ", "_")
#             
#             bioproject_phenodata_dir = phenodata_dir + sheet_title
#             if not os.path.exists(bioproject_phenodata_dir):
#                 os.makedirs(bioproject_phenodata_dir)
#             
#             # Write out the tsv
#             fd = open(bioproject_phenodata_dir + "/phenodata.tsv", "w", encoding="utf-8")
#             writer = csv.writer(fd, delimiter='\t')
#             row_num = 0
#             for row in wsheet.iter_rows():
#                 
#                 row_data = []
#                 for cell in row:
#                     if cell.value is not None:
#                         row_data.append(cell.value)
#                         
#                 if row_num == 0:
#                     row_data = [x.strip().replace(" ", ".") for x in row_data]
#                     
#                 row_num += 1
#                 
#                 writer.writerow(row_data)
#             
#             fd.close()
            
    response = {"message": "Phenodata file correctly saved.", "type": "info"}
    
    return HttpResponse(json.dumps(response))


def upload_data(request, project_id):
    
    print(request)
    
    response = []
    
    total = 0
    for file in request.FILES.values():
        file_dir = get_file_dir(project_id)
        
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
#         
        filepath = file_dir + file.name
        default_storage.save(filepath, ContentFile(file.read()))
        
        response.append({"message": {"url": make_relative(filepath), "name": file.name}, "type": "data"})
        total += 1
        
    response.append({"message": str(total) + " files(s) correctly saved.", "type": "info"})
    
    return HttpResponse(json.dumps(response))


def upload_temp(request):
    
    print(request)
    
    response = []
    
    total = 0
    for file in request.FILES.values():
        file_dir = get_temp_dir()
        
        if not os.path.exists(file_dir):
            os.makedirs(file_dir)
#         
        filepath = file_dir + file.name
        default_storage.save(filepath, ContentFile(file.read()))
        
        response.append({"message": {"url": make_relative(filepath), "name": file.name}, "type": "data"})
        total += 1
        
    response.append({"message": str(total) + " temporary files(s) correctly saved.", "type": "info"})
    
    return HttpResponse(json.dumps(response))


def upload_pipeline_data(request, project_id, pipeline_id):
    
    print(request)
    
    response = []
    
    total = 0
    for file in request.FILES.values():
        data_dir = get_pipeline_script_dir(project_id, pipeline_id)
        
        if not os.path.exists(data_dir):
            os.makedirs(data_dir)
#         
        filepath = data_dir + file.name
        path = default_storage.save(filepath, ContentFile(file.read()))
        st = os.stat(path)
        os.chmod(filepath, st.st_mode | stat.S_IEXEC)
        
        url = "/data/" + project_id + "/scripts/" + pipeline_id + "/" + file.name
        response.append({"message": {"url": url, "name": file.name}, "type": "data"})
        total += 1
        
    response.append({"message": str(total) + " files(s) correctly saved.", "type": "info"})
    
    return HttpResponse(json.dumps(response))


def remove_data(request, project_id):
    
    print(request)
    file = json.loads(request.body.decode('utf-8'))["file"]
    response = [{"message": "File " + file["name"] + " correctly removed.", "type": "info"}]
    
    print(request.FILES)
    data_dir = get_project_dir(project_id)
    filepath = data_dir + file["name"]
    path = default_storage.delete(filepath)
        
    return HttpResponse(json.dumps(response))


def remove_pipeline_data(request, project_id, pipeline_id):
    
    print(request)
    file = json.loads(request.body.decode('utf-8'))["file"]
    response = [{"message": "File " + file["name"] + " correctly removed.", "type": "info"}]
    
    print(request.FILES)
    data_dir = get_project_dir(project_id) + "scripts/" + pipeline_id + "/"
    filepath = data_dir + "/" + file["name"]
    path = default_storage.delete(filepath)
        
    return HttpResponse(json.dumps(response))


def remove_run(request, project_id, run_id):
    
    response = []
    
    project = load_project_from_disk(project_id)
    deleted_run = None
    for run in project["runs"]:
        if run["id"] == run_id:
            deleted_run = run
            project["runs"].remove(run)
            break
    
    local_removal_error = False
    run_dir = get_run_dir(deleted_run)
    print("Going to remove directory", run_dir)
    try:
        shutil.rmtree(run_dir)
    except FileNotFoundError:
        pass
    except Exception as e:
        print("EXCEPTION", type(e), e)
        local_removal_error = True
    
    pipeline = deleted_run["pipelines"][0]
    real_run_dir = get_real_run_dir(pipeline, deleted_run)
    print("Going to remove directory", real_run_dir)
    
    username, remote_path, hostname = get_credentials(pipeline)
    remote_command = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                            "bash -c 'cd " + real_run_dir + "; cat job_ids | xargs -n 1 scancel; cd " + remote_path + "; find " + remote_path + " -name \"*" + deleted_run["id"] + "*\" -print0 | xargs -0 -r rm -r ';"])
    if remote_command["exit_code"] != 0:
        response.append(remote_command)
    
    if not local_removal_error and remote_command["exit_code"] == 0:
        save_project_raw(project)
        response.append({"message": "Run " + run_id + " correctly removed.", "exit_code": 0, "type": "info"})
    else:
        response.append({"message": "There was some problem during removal of run " + run_id, "type": "warning"})
    
    return HttpResponse(json.dumps(response))


def stop_jobs(request, project_id, run_id):
    
    response = []
    
    project = load_project_from_disk(project_id)
    deleted_run = None
    for run in project["runs"]:
        if run["id"] == run_id:
            deleted_run = run
            break
    
    pipeline = deleted_run["pipelines"][0]
    real_run_dir = get_real_run_dir(pipeline, deleted_run)
    
    username, remote_path, hostname = get_credentials(pipeline)
    remote_command = execute_remote_command(["ssh", "-i", "pipeline_id_rsa", "-oPasswordAuthentication=no", username + "@" + hostname,
                            "bash -c 'cd " + real_run_dir + "; cat job_ids | xargs -n 1 scancel;'"])
    if remote_command["exit_code"] != 0:
        response.append(remote_command)
    
    if remote_command["exit_code"] == 0:
        save_project_raw(project)
        response.append({"message": "Jobs of run " + run_id + " correctly stopped.", "exit_code": 0, "type": "info"})
    else:
        response.append({"message": "There was some problem while stopping jobs of run " + run_id, "type": "warning"})
    
    return HttpResponse(json.dumps(response))


from openpyxl import Workbook
from wsgiref.util import FileWrapper


def create_phenodata(project, bioprojects=None, mode="separate"):
    wb = Workbook()
    wb.remove(wb.active)
    
    attribute_keys = set()
    for bioproject in project["projects"]:
        if bioprojects and bioproject["id"] not in bioprojects: continue
        
        for exp in bioproject["experiments"]:
            if "attributes" in exp["dataset"]:
                for attribute_key in exp["dataset"]["attributes"].keys():
                    attribute_keys.add(attribute_key)
                    
            for tag in exp["tags"]:
                type = tag["type"]
                
                if type in ["Type", "Experiment"]: continue
                attribute_keys.add(type)
    
    attribute_keys = ["ids"] + list(attribute_keys)
    
    if mode == "joined":
        ws = wb.create_sheet(project["id"])
        for j, h in enumerate([a.replace(" ", "_") for a in attribute_keys]):
            ws.cell(row=1, column=j + 1, value=h)
    
    n = 0
    for bioproject in project["projects"]:
        if bioprojects and bioproject["id"] not in bioprojects: continue
        
        bioproject_id = bioproject["id"]
        if mode != "joined":
            ws = wb.create_sheet(bioproject_id)
            for j, h in enumerate([a.replace(" ", "_") for a in attribute_keys]):
                ws.cell(row=1, column=j + 1, value=h)
        
#         attribute_keys_to_remove = []
        if "attributes" in exp["dataset"]:
            for attribute_key in attribute_keys:
                attribute_values = Counter()
                for exp in bioproject["experiments"]:
                    if "attributes" in exp["dataset"]:
                        if attribute_key in exp["dataset"]["attributes"]:
                            value = exp["dataset"]["attributes"][attribute_key]
                            attribute_values[value] += 1

#             # Remove columns which have always the same value             
#             if len(attribute_values.keys()) == 1:
#                 print("Discarding column {} for ProjectID={} (not discriminating) Total={} Counter={}".format(attribute_key, bioproject_id, len(bioproject["experiments"]), attribute_values))
#                 attribute_keys_to_remove.append(attribute_key)
#         
#         for attribute_key in attribute_keys_to_remove:
#             attribute_keys.remove(attribute_key)
        
        if mode != "joined":
            n = 0
                
        for i, exp in enumerate(sorted(bioproject["experiments"], key=lambda x: x["id"])):
            for j, attribute_key in enumerate(attribute_keys):
                value = ""
                if attribute_key == "ids":
                    value = exp["id"]

                if "attributes" in exp["dataset"]:
                    if attribute_key in exp["dataset"]["attributes"]:
                        value = exp["dataset"]["attributes"][attribute_key]
                        
                if not value:
                    value = []
                    for tag in exp["tags"]:
                        if tag["type"] == attribute_key:
                            value.append(tag["name"])
                            
                    value = ",".join(value)
                
                ws.cell(row=n + 2, column=j + 1, value=value)
            
            n += 1
    
    return wb


def download_phenodata(request):
    data = json.loads(request.body.decode('utf-8'))
    
    project = data["project"]
    choices = data["options"]

    print(choices)
    
    filename = "phenodata.xlsx"
    
    phenodata_dir = get_project_dir(project["id"])
    filepath = phenodata_dir + "/" + filename
    
    if os.path.exists(filepath):
        print("PHENODATA FILEPATH", filepath)
        with open(filepath, 'rb') as f:
           file_data = f.read()
           
           response = HttpResponse(file_data, content_type='application/ms-excel')
    else:
        response = HttpResponse(content_type='application/ms-excel')
        wb = create_phenodata(project, mode=choices[0]["choice"])
        wb.save(response)
        
    response['Content-Disposition'] = 'attachment; filename=' + filename
    
    return response


def get_phenodata(request, project_id, bioproject_id):
    header = []
    rows = []
    
    project = load_project_from_disk(project_id)
    wb = create_phenodata(project, [bioproject_id])
    for wsheet in wb.worksheets:
        line_no = 1
        for row in wsheet.iter_rows():
            table_row = []
            
            for cell in row:
                table_row.append(str(cell.value))
                
            if line_no == 1:
                header = [x for x in table_row]
            else:
                rows.append(table_row)
            
            line_no += 1
    
    result = {"title": bioproject_id, "header": header, "rows": rows}
    
    return HttpResponse(json.dumps(result))


def create_dataset(project):
    wb = Workbook(write_only=True)
    
    header = ["run", "experiment", "bioproject", "path", "path2"]
    
    for bioproject in project["projects"]:
        
        ws = wb.create_sheet(bioproject["id"])
        ws.append(header)
        
        for experiment in bioproject["experiments"]:
            for sample in experiment["dataset"]["sample_ids"]:
                values = [sample["id"], experiment["id"], bioproject["id"]]
                
                if "paths" in sample:
                    values.append(sample["paths"][0])
                    if len(sample["paths"]) > 1:
                        values.append(sample["paths"][1])
                
                ws.append(values)
    return wb


def download_dataset(request):
    project = json.loads(request.body.decode('utf-8'))
    
    filename = "dataset.xlsx"
    
    dataset_dir = get_project_dir(project["id"])
    filepath = dataset_dir + "/" + filename
    
#     if os.path.exists(filepath):
#         print(filepath)
#         with open(filepath, 'rb') as f:
#            file_data = f.read()
#            
#            response = HttpResponse(file_data, content_type='application/ms-excel')
#     else:
    response = HttpResponse(content_type='application/ms-excel')
    wb = create_dataset(project)
    wb.save(response)
        
    response['Content-Disposition'] = 'attachment; filename=' + filename
    
    return response
